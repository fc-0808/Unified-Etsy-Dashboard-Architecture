"""
generate_shopping_route.py

Parses Etsy order PDFs, cross-references each item against the supplier
catalog, and produces ``shopping_route.xlsx`` -- a purchasing guide
organised by supplier location for efficient in-person shopping.

Also updates ``supplier_catalog.xlsx`` with any products not yet catalogued
(adds them with their product photo so you can fill in supplier info later).
A duplicate guard prevents the same title from being appended more than once
across multiple runs.  The Shopping Route places such items in a dedicated
blue "Awaiting Supplier Info" section so they are never buried in the table.

**Charms (two layers):**

1. **Charm Library** sheet — master list of physical charm designs: **A** photo, **B** stable
   code, **C** SKU (sortable label), **D** default charm shop.  One row per distinct charm
   you stock.  Disk assets ``data/charm_images/<Code>.png`` (etc.) override embedded
   photos when generating routes.  New codes use zero-padded suffixes (e.g. ``CH-00001``;
   width follows existing rows/files).

2. **Product Map** — each product row is a phone case / listing variant.  **Column F**
   (*Charm Shop*) = which stall supplies the charm for *that product*.  **Column G**
   (*Charm Code*) = optional link to **Charm Library** column **B**.  **Column H** = *NOTES*.

   **Discontinuing a product:** use the UI button or ``--mark-product-discontinued``.
   The row is **moved** from Product Map to the **Discontinued Products** sheet
   (with a timestamp and photo), so the catalog stays clean. The duplicate guard
   checks both sheets, preventing re-addition.

   **When column G (Charm Code) is set** to a library code, the shopping route (and HTML) **Charm**
   section **aggregates by charm code** — one row per unique charm with the library photo,
   code, SKU, charm shop, and total quantity across all orders.  Folder file wins over embed.

   **When Charm Code is blank**, the order is shown in a separate **"Awaiting Charm Code"**
   sub-section with the product photo and a prompt to assign a code in the catalog.

   **Charm section vs dashboard:** Only line items whose charm Buy Status is still actionable
   (Pending or Out of Stock — not Purchased or Out of Production) appear in the route Charm
   section, so the spreadsheet lists what you still need to buy.  Line items you uncheck in
   the dashboard are removed from *all* route sections (including charm aggregation) via
   ``--exclude-orders-file``.

   Typical setup: add every distinct charm to the library once; for each catalog product
   that ships with that charm, pick the matching **G** (shop) and **H** (code).  **G** can
   match **D** from the library as a default, but you may override **G** per product if
   sourcing differs.

Unless ``--no-charm-manifest``, each run writes ``data/charm_manifest.json``; use
``--export-charm-manifest`` for that step only.

Usage
-----
    python generate_shopping_route.py                         # auto-discover PDFs
    python generate_shopping_route.py order1.pdf order2.pdf   # explicit PDFs
    python generate_shopping_route.py --threshold 60          # lower match bar
    python generate_shopping_route.py --no-catalog-update     # skip catalog write

Adding a new batch of today's orders
-------------------------------------
Each day you download new order PDFs, run:

    python generate_shopping_route.py --new-batch

The script auto-discovers all *.pdf files in the current directory.  Any PDF
whose filename was already processed in a previous run is automatically
skipped (its orders are already in the cache).  Only brand-new PDFs are
parsed and merged on top of the existing shopping route.  After processing,
every ingested PDF filename is recorded in orders_cache.json so it is never
re-parsed on future runs.

If you delete ``shopping_route.xlsx`` from the output folder (or it is absent)
and run again **without** ``--refresh-catalog``, prior orders are **not** read
from ``orders_cache.json`` or any leftover path: the script rebuilds solely
from the PDFs currently in ``input/`` (same idea as ``--reset``).  Use
``--refresh-catalog`` when you need to regenerate the Excel from the cache
after deleting only the output file.

You may also pass the PDF files explicitly instead of auto-discovery:

    python generate_shopping_route.py --new-batch file1.pdf file2.pdf ...

Post-shopping cleanup
---------------------
After returning from a shopping trip and updating the status dropdowns in
shopping_route.xlsx, run:

    python generate_shopping_route.py --purge-purchased

This reads every status you have already entered in the Excel file and applies
**independent section-level purging**:

  • Case / Grip section  (supplier floors) and
  • Charm section        (separate building)

are evaluated separately.  If the Case/Grip section is fully purchased (every
present Case/Grip component is "Purchased") it is stripped from the item's
style even if the Charm is still pending -- and vice versa.  An item is only
removed from the route entirely when *both* sections are fully purchased.

Items with any component still "Out of Stock" or "Pending" are kept.
"Out of Production" is treated as complete (same as Purchased); those items
are purged and recorded to out_of_production_log.csv so they never reappear.
The cache (orders_cache.json) is updated to match, so partially-purged items
re-appear on the next run with only their remaining section.
"""
from __future__ import annotations
import argparse
import base64
import csv
import fnmatch
import json
import logging
import os
import re
import secrets
import shutil
import sqlite3
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
import fitz          # pymupdf -- for JPEG extraction
import pdfplumber
import openpyxl
try:
    from deep_translator import GoogleTranslator as _GoogleTranslator
    _DEEP_TRANSLATOR_AVAILABLE = True
except ImportError:
    _DEEP_TRANSLATOR_AVAILABLE = False
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import EMU_to_pixels, pixels_to_EMU, points_to_pixels
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from rapidfuzz import fuzz, process
from supplier_catalog_backup import (
    backup_supplier_catalog_before_write,
    catalog_backup_dir,
    list_supplier_catalog_backups,
    restore_supplier_catalog,
)
CATALOG_FILE      = "supplier_catalog.xlsx"
CATALOG_SHEET     = "Product Map"
SUPPLIERS_SHEET   = "Suppliers"
DISCONTINUED_SHEET = "Discontinued Products"
CHARM_SHOPS_SHEET = "Charm Shops"          # separate sheet for charm-shop reference data
CHARM_LIBRARY_SHEET = "Charm Library"    # one row per physical charm (shared photo + optional default shop)
CHARM_IMAGES_DIR_NAME = "charm_images"
CHARM_CODES_LIST_MAX_ROW = 10000
CHARM_CODE_NUMERIC_MIN_WIDTH = 5
CHARM_LIBRARY_COL_SKU_HEADER = "SKU"
CHARM_LIBRARY_INSTRUCTION_MARKER = "charm library — quick guide"
CHARM_LIBRARY_INSTRUCTION_TEXT = (
    "Charm Library — quick guide\n\n"
    "• One charm per row.  Column B = stable code (e.g. CH-00001);  "
    "Column C = SKU / short label;  Column D = Default Charm Shop.\n\n"
    "• Photo: column A, or file data/charm_images/<Code>.png  "
    "(folder file takes priority when both exist).\n\n"
    "• Product Map → column F = Charm Shop (where to buy);  "
    "column G = Charm Code (same as column B here).  "
    "Leave G blank to use the order PDF photo.\n\n"
    "• To insert a new charm between existing ones:\n"
    "  1. Import the photo via the Python UI (it appends at the end).\n"
    "  2. Cut its row here and paste it where it belongs "
    "(right-click a row number → Insert Cut Cells).\n"
    "  3. Open the Python UI → Tab 2 → Section C → click "
    "\"Renumber charm codes\" — codes are reassigned by row order "
    "and Product Map references update automatically.\n\n"
    "• After editing: save this workbook, then run the route generator "
    "(--refresh-catalog)."
)
_CHARM_LIBRARY_INSTRUCTION_VERSION = "v2-reorder"
CHARM_SHOPS_INSTRUCTION_MARKER = "charm shops — quick guide"
CHARM_SHOPS_INSTRUCTION_TEXT = (
    "Charm Shops — quick guide\n\n"
    "• Edit the rows above (shop name + stall).\n\n"
    "• Save the workbook, then run the route generator."
)
OUTPUT_FILE       = "shopping_route.xlsx"
CHARM_MANIFEST_FILE = "charm_manifest.json"   # default under data/ with --project-dir
CACHE_FILE        = "orders_cache.json"
DB_FILE           = "etsy_orders.db"         # SQLite DB (replaces xlsx catalog reads + JSON cache)
OOP_LOG_FILE      = "out_of_production_log.csv"   # append-only log of purged OOP items
ZH_TRANS_CACHE    = "translations_zh_cache.json"   # persisted product-title translations
MATCH_THRESHOLD   = 65
_DEFAULT_CHARM_SHOPS = [
    ("彩虹",     "2D21",    ""),
    ("有米UMI",  "2D02",    ""),
    ("長金飾品", "2D04",    ""),
    ("一樂潮品", "2C666",   ""),
    ("小艾飾品", "2D41-43", ""),
]
EMPTY_ENTRY_MATCH_THRESHOLD  = 90
FILLED_ENTRY_MATCH_THRESHOLD = 85
SAME_PRODUCT_THRESHOLD = 97
VARIANT_HINT_THRESHOLD = SAME_PRODUCT_THRESHOLD
PHOTO_PX    = 155   # square thumbnail px (enlarged for better visibility)
ROW_HEIGHT  = 120.0 # Excel row height in points (~160 px at 96 dpi)
PHOTO_COL_W = 26.0  # Excel column width in character units
ZH_ROW_HEIGHT  = 210.0  # ~280 px at 96 dpi
ZH_PHOTO_PX    = 265    # square thumbnail px for ZH sheet
ZH_PHOTO_COL_W = 44.0   # must cover ZH_PHOTO_PX — was 34 and caused overlap into col C
CHARM_LIB_ROW_HEIGHT    = ZH_ROW_HEIGHT
CHARM_LIB_COL_A_WIDTH   = ZH_PHOTO_COL_W
CHARM_LIB_CELL_PAD_PX   = 4   # small inset from cell edges (gridlines / anti-aliasing)
_OPENAI_DEFAULT_BASE = "https://api.openai.com/v1"
_CHARM_VISION_COOLDOWN_SEC = 0.35   # light spacing between API calls on bulk import
_CHARM_VISION_API_KEY_ENVS = (
    "CHARM_VISION_API_KEY",
    "OPENAI_API_KEY",
    "OPENROUTER_API_KEY",
)
_CHARM_VISION_BASE_URL_ENVS = ("CHARM_VISION_BASE_URL", "OPENAI_BASE_URL")
logging.basicConfig(level=logging.INFO, format="%(levelname)-8s  %(message)s")
log = logging.getLogger("shopping_route")
STATUS_OPTIONS = ["Pending", "Purchased", "Out of Stock", "Out of Production"]
ZH_STATUS_OPTIONS = ["待处理", "已购买", "缺货", "停产"]
_ZH: dict[str, str] = {
    # Sheet / tab names
    "Shopping Route":   "购物路线",
    "Orders Detail":    "订单明细",
    "Summary":          "汇总",
    # Route sheet column headers
    "Photo":            "图片",
    "Floor":            "楼层",
    "Supplier":         "供应商",
    "Stall":            "摊位",
    "Product":          "产品",
    "Items to Purchase": "待购项",
    "Case":             "手机壳",
    "Grip":             "支架",
    "Charm":            "挂件",
    "Phone Model":      "手机型号",
    "Qty":              "数量",
    "Order #":          "订单号",
    "Etsy Shop":        "Etsy店铺",
    # Orders Detail extra headers
    "Buyer":            "买家",
    "Ship To":          "收货人",
    "Country":          "国家",
    "Order Date":       "订单日期",
    "Match %":          "匹配度",
    # Summary stat labels
    "Total orders":                          "订单总数",
    "Total line items":                      "商品行数",
    "Total quantity":                        "总数量",
    "Ready (supplier + location)":           "就绪（供应商+位置齐全）",
    "In catalog \u2013 needs supplier info": "目录中\u2013待补供应商信息",
    "Not in catalog (unmatched)":            "不在目录中（未匹配）",
    "Items per Supplier":                    "各供应商商品数",
    "Items per Etsy Shop":                   "各Etsy店铺商品数",
    "Items":                                 "商品数",
    "Orders":                                "订单数",
    # Status cell values (dropdown + written into cells)
    "Pending":           "待处理",
    "Purchased":         "已购买",
    "Out of Stock":      "缺货",
    "Out of Production": "停产",
    "N/A":               "不适用",
    "case only":         "仅手机壳",
    "grip only":         "仅支架",
    "case, grip":        "手机壳、支架",
    # Charm section strings
    "Charms to Purchase":    "待购挂件",
    "Separate Building":     "独立楼栋",
    "charm(s) needed across": "个挂件，涉及",
    "order(s)":              "个订单",
    "Charm shops":           "挂件店铺",
    "No charm shops configured": "未配置挂件店铺",
    "Private Notes":         "私信备注",
    # Simple-route column headers (not already covered above)
    "Charm Code":            "挂件编码",
    "Charm Shop":            "挂件店铺",
    # Simple-route subtitle fragments
    "sorted lowest to highest floor": "按楼层从低到高排序",
    "awaiting supplier info":         "待填供应商信息",
    "unmatched":                      "未匹配",
    "separate building":              "独立楼栋",
    # Status legend row
    "Status legend: Pending (white) | Purchased (green) | Out of Stock (amber) | Out of Production (red) | N/A (gray)":
        "状态说明：  待处理（白）|  已购买（绿）|  缺货（黄）|  停产（红）|  不适用（灰）",
    # Awaiting / needs-info banner phrases
    "In Catalog \u2013 Awaiting Supplier Info":
        "目录中 \u2013 待填供应商信息",
    "open supplier_catalog.xlsx and fill in Shop Name + Stall":
        "\u8bf7\u5728 supplier_catalog.xlsx \u4e2d\u586b\u5199\u300c\u5e97\u94fa\u540d\u300d\u548c\u300c\u644a\u4f4d\u300d\u5217",
    "Unmatched Items \u2014 supplier not found in catalog":
        "未匹配商品 \u2014 目录中未找到供应商",
    # Charm banner
    "CHARMS TO PURCHASE \u2014 SEPARATE BUILDING": "待购挂件 \u2014 独立楼栋",
    "charm(s) needed across %d order(s)": "个挂件，涉及 %d 个订单",
    "order(s) missing charm-code assignment": "个订单缺挂件编码",
    "order(s) missing charm-shop assignment": "个订单缺挂件店铺",
    "No charm shops configured \u2014 add them in the \u2018Charm Shops\u2019 tab":
        "\u672a\u914d\u7f6e\u6302\u4ef6\u5e97\u94fa \u2014 \u8bf7\u5728\u300c\u6302\u4ef6\u5e97\u94fa\u300d\u5de5\u4f5c\u8868\u4e2d\u6dfb\u52a0",
    # Awaiting charm code sub-section banner
    "AWAITING CHARM CODE ASSIGNMENT": "待分配挂件编码",
    "open supplier_catalog.xlsx \u2192 Product Map col H (Charm Code)":
        "请在 supplier_catalog.xlsx \u2192 产品映射 H 列（挂件编码）中填写",
    "\u23f3 Awaiting Code": "\u23f3 待分配编码",
}
def _t(key: str, lang: str) -> str:
    """Return the Simplified-Chinese translation of *key* when lang=='zh',
    otherwise return the key unchanged."""
    return _ZH.get(key, key) if lang == "zh" else key
_STATUS_FILLS = {
    "Purchased":         PatternFill("solid", fgColor="C6EFCE"),  # green
    "Out of Stock":      PatternFill("solid", fgColor="FFEB9C"),  # amber
    "Out of Production": PatternFill("solid", fgColor="FFC7CE"),  # red
}
_STATUS_FONTS = {
    "Purchased":         Font("Calibri", size=10, color="276221"),
    "Out of Stock":      Font("Calibri", size=10, color="7D4E00"),
    "Out of Production": Font("Calibri", size=10, color="9C0006"),
}
_ITEMS_TO_PURCHASE_FONT = Font("Calibri", size=10, bold=True)
@dataclass
class OrderItem:
    title:       str
    quantity:    int   = 1
    phone_model: str   = ""
    style:       str   = ""
    photo_bytes: bytes | None = None   # raw JPEG bytes extracted from PDF
    # Etsy listing id supplied by the Unified Dashboard (--import-json).  This is
    # the ONLY stable per-line discriminator: two genuinely-different listings on
    # the same order can share the first 50 normalised title characters AND the
    # same component set, so keying line dedup on title alone collapses them into
    # one and silently drops a product the operator still has to buy.  Folding
    # the listing id into the dedup key keeps such lines distinct.  Empty for
    # PDF-parsed orders (no listing id) and manual items — the dedup then behaves
    # exactly as before (title-only), so nothing regresses.
    listing_id:  str = ""
    # Optional manual charm assignment supplied by the Unified Dashboard
    # (Route tab).  When set, these override whatever the catalog match would
    # have produced for this line — see _apply_import_charm_overrides().
    charm_code_override: str = ""
    charm_shop_override: str = ""
    # Optional supplier (shop + stall) resolved by the Unified Dashboard.  The
    # dashboard is the source of truth: it already shows the operator the exact
    # shop/stall (from catalog enrichment + manual overrides), so we carry those
    # values through the import and stamp them onto the resolved supplier — see
    # _apply_import_supplier_overrides().  This guarantees the route matches the
    # dashboard even for products OSP's own catalog hasn't fully mapped yet.
    supplier_shop_override: str = ""
    supplier_stall_override: str = ""
@dataclass
class Order:
    order_number:   str = ""
    etsy_shop:      str = ""
    buyer_name:     str = ""
    buyer_username: str = ""
    ship_to_name:   str = ""
    ship_to_country: str = ""
    order_date:     str = ""
    private_notes:  str = ""
    items: list[OrderItem] = field(default_factory=list)
@dataclass
class CatalogEntry:
    product_title: str = ""
    category:      str = ""
    shop_name:     str = ""
    stall:         str = ""
    price:         str = ""
    notes:         str = ""
    # Foreign key → Charm Shops tab: which charm shop supplies the charm for
    # this product.  Empty string means no charm, or charm shop not yet assigned.
    charm_shop:    str = ""
    # Foreign key → Charm Library tab: stable code for the physical charm (shared photo).
    charm_code:    str = ""
    # True when the shop/stall came from an authoritative source (the Unified
    # Dashboard, where the operator has confirmed the supplier) rather than a
    # fuzzy catalog match.  Authoritative suppliers bypass the match-confidence
    # heuristic in _needs_catalog_entry() so they are always routed to their
    # known location instead of being dumped into the "unmatched / ???" section.
    authoritative: bool = False
@dataclass
class CharmLibraryEntry:
    """One row in the Charm Library sheet — a reusable charm SKU."""
    code:               str              # column B, unique key
    sku:                str = ""         # column C — short label / stock SKU
    default_charm_shop: str = ""         # optional; same names as Charm Shops col A
    notes:              str = ""
    photo_bytes:        bytes | None = None
@dataclass
class CharmShop:
    """One entry in the Charm Shops reference sheet."""
    shop_name: str = ""
    stall:     str = ""
    notes:     str = ""
@dataclass
class ResolvedItem:
    order:       Order
    item:        OrderItem
    supplier:    CatalogEntry | None = None
    match_score: float = 0.0
_COL_BOUNDARY     = 200   # px dividing left address block from right product block
_LEFT_META_MAX_X  = 100   # px — order metadata (order#, address, shop, date) lives at x≈36
_LEFT_PN_MIN_X    = 100   # px — Private notes content lives at x≈174 (separate sub-column)
_LEFT_PN_MAX_Y    = 500   # px — crop footer ("Do the green thing", etc. at y≈700+)
_Y_TOLERANCE      = 3.0   # px -- merge words within this gap into one line
_QUANTITY_RE          = re.compile(r"^Quantity:\s*(\d+)")
_MODEL_RE             = re.compile(r"^(?:Phone|iPhone)\s+Model:\s*(.+)", re.IGNORECASE)
_STYLE_RE             = re.compile(r"^Styles?:\s*(.+)")
_CURRENCY_RE          = re.compile(r"^[A-Z]{3}$")
_ORDER_RE             = re.compile(r"^Order\s+#(\d+)")
_SCHEDULED_RE         = re.compile(r"^Scheduled\s+to\s+(?:ship|dispatch)\s+by", re.IGNORECASE)
_PRIVATE_NOTES_RE     = re.compile(r"^Private\s+notes?$", re.IGNORECASE)
_TRAILING_CURRENCY_RE = re.compile(r"\s+[A-Z]{3}$")
def import_orders_from_json(path: Path) -> list[Order]:
    """Load orders from a Unified Dashboard JSON export file.

    Replaces PDF parsing when the Unified Etsy Dashboard is running and has
    already synced orders from Etsy directly.  Call the dashboard export
    endpoint (``GET /api/export/orders-for-route``) to produce the file,
    then pass it here via ``--import-json``.

    The JSON schema mirrors what the dashboard endpoint returns::

        {
            "orders": [
                {
                    "order_number":    "1234567890",
                    "etsy_shop":       "MyShopName",
                    "buyer_name":      "Jane Smith",
                    "buyer_username":  "janesmith",
                    "ship_to_name":    "Jane Smith",
                    "ship_to_country": "US",
                    "order_date":      "Jan 15, 2025",
                    "private_notes":   "Please gift wrap",
                    "items": [
                        {
                            "title":       "Cute Bunny Case + Grip Set",
                            "quantity":    1,
                            "phone_model": "iPhone 15 Pro Max",
                            "style":       "Case+Grip",
                            "image_url":   "https://...",
                            "charm_code":  "CH-00007",   # optional manual override
                            "charm_shop":  "彩虹"          # optional manual override
                        }
                    ]
                }
            ]
        }

    ``photo_bytes`` is left as ``None``; the catalog-photo replacement step
    (``apply_catalog_photos_to_resolved``) fills in photos for matched items.
    """
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        log.error("Cannot read --import-json file %s: %s", path, exc)
        sys.exit(1)

    raw_orders = raw.get("orders", raw) if isinstance(raw, dict) else raw
    if not isinstance(raw_orders, list):
        log.error("--import-json: expected a list of orders (or {orders: [...]}) in %s", path)
        sys.exit(1)

    orders: list[Order] = []
    for raw_order in raw_orders:
        if not isinstance(raw_order, dict):
            continue

        raw_items = raw_order.get("items", [])
        if not raw_items:
            continue

        order = Order(
            order_number    = str(raw_order.get("order_number", "")),
            etsy_shop       = str(raw_order.get("etsy_shop", "")),
            buyer_name      = str(raw_order.get("buyer_name", "")),
            buyer_username  = str(raw_order.get("buyer_username", "")),
            ship_to_name    = str(raw_order.get("ship_to_name", "")),
            ship_to_country = str(raw_order.get("ship_to_country", "")),
            order_date      = str(raw_order.get("order_date", "")),
            private_notes   = str(raw_order.get("private_notes", "")),
        )

        for raw_item in raw_items:
            if not isinstance(raw_item, dict):
                continue
            title = str(raw_item.get("title", "")).strip()
            if not title:
                continue
            # photo_bytes: the Unified Dashboard embeds the Etsy CDN image as a
            # base64 string ("image_b64") so the Excel route files always show
            # the product photo even for items not yet in the supplier catalog.
            # apply_catalog_photos_to_resolved (called later) overrides this
            # with the canonical Product Map photo for matched items — keeping
            # the well-curated catalog photo as the definitive source of truth.
            _image_b64 = str(raw_item.get("image_b64", "") or "").strip()
            _photo: bytes | None = None
            if _image_b64:
                try:
                    _photo = base64.b64decode(_image_b64)
                except Exception:
                    pass   # malformed base64 — fall back to catalog photo only

            order.items.append(OrderItem(
                title       = title,
                quantity    = int(raw_item.get("quantity", 1) or 1),
                phone_model = str(raw_item.get("phone_model", "")),
                style       = str(raw_item.get("style", "")),
                photo_bytes = _photo,
                listing_id  = str(raw_item.get("listing_id", "") or "").strip(),
                charm_code_override = str(raw_item.get("charm_code", "") or "").strip(),
                charm_shop_override = str(raw_item.get("charm_shop", "") or "").strip(),
                supplier_shop_override  = str(raw_item.get("supplier_shop", "") or "").strip(),
                supplier_stall_override = str(raw_item.get("supplier_stall", "") or "").strip(),
            ))

        if order.items:
            orders.append(order)

    log.info(
        "import_orders_from_json: loaded %d order(s), %d line item(s) from %s",
        len(orders),
        sum(len(o.items) for o in orders),
        path.name,
    )
    return orders
def _apply_import_charm_overrides(resolved: list["ResolvedItem"]) -> int:
    """Apply manual charm assignments carried on imported OrderItems.

    The Unified Dashboard's Route tab lets the operator assign a specific charm
    (``charm_code`` + ``charm_shop``) to an order line.  Those values ride along
    on ``OrderItem.charm_code_override`` / ``charm_shop_override`` through the
    JSON import.  After the automatic catalog match has run we stamp them onto
    each item's resolved supplier entry so they win over whatever the fuzzy
    match produced (or supply a charm where the catalog had none).

    Returns the number of line items whose charm assignment was overridden.
    """
    count = 0
    for r in resolved:
        code = getattr(r.item, "charm_code_override", "") or ""
        shop = getattr(r.item, "charm_shop_override", "") or ""
        if not code and not shop:
            continue
        # CRITICAL: match_items assigns the SAME CatalogEntry object to every
        # item that matched a given catalog row.  Mutating it in place would
        # corrupt every other item sharing it (last-write-wins across unrelated
        # orders).  Always attach a private copy before applying overrides.
        if r.supplier is None:
            r.supplier = CatalogEntry(product_title=r.item.title)
        else:
            r.supplier = replace(r.supplier)
        if code:
            r.supplier.charm_code = code
        if shop:
            r.supplier.charm_shop = shop
        count += 1
    if count:
        log.info("Applied %d manual charm override(s) from dashboard import", count)
    return count
def _apply_import_supplier_overrides(resolved: list["ResolvedItem"]) -> int:
    """Apply supplier (shop + stall) resolved by the Unified Dashboard.

    The dashboard is the authoritative source for which physical shop / stall an
    order line maps to — it merges the supplier catalog match with the
    operator's manual per-product and per-order overrides.  Those resolved
    values ride along on ``OrderItem.supplier_shop_override`` /
    ``supplier_stall_override`` through the JSON import.

    After the automatic catalog match has run we stamp them onto each item's
    resolved supplier entry so the route's floor/stall placement always agrees
    with what the operator sees in the dashboard — even for products that OSP's
    own ``supplier_catalog.xlsx`` has not (yet) mapped.  Charm fields set by the
    catalog match or by ``_apply_import_charm_overrides`` are preserved.

    Returns the number of line items whose supplier was overridden.
    """
    count = 0
    for r in resolved:
        shop  = getattr(r.item, "supplier_shop_override", "") or ""
        stall = getattr(r.item, "supplier_stall_override", "") or ""
        if not shop and not stall:
            continue
        # CRITICAL: detach from the shared catalog object before mutating (see
        # _apply_import_charm_overrides) so overriding one order's supplier never
        # bleeds into a different product that fuzzy-matched the same catalog row.
        if r.supplier is None:
            r.supplier = CatalogEntry(product_title=r.item.title)
        else:
            r.supplier = replace(r.supplier)
        if shop:
            r.supplier.shop_name = shop
        if stall:
            r.supplier.stall = stall
        # Mark as operator-confirmed so routing trusts it regardless of the
        # automatic match score (see _needs_catalog_entry).  Catalog-completeness
        # (amber "new product" rows) is unaffected — that uses _needs_own_catalog_row.
        r.supplier.authoritative = True
        count += 1
    if count:
        log.info("Applied %d supplier override(s) from dashboard import", count)
    return count
def get_db_connection(db_path: Path) -> sqlite3.Connection:
    """Open a WAL-mode SQLite connection with Row factory and FK enforcement.

    Use as a context manager or close explicitly when done:
        conn = get_db_connection(db_path)
        ...
        conn.close()
    """
    conn = sqlite3.connect(str(db_path), timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    # Wait (rather than fail immediately) when another connection briefly holds
    # the write lock — the dashboard keeps a read-only handle on this same file.
    conn.execute("PRAGMA busy_timeout=30000")
    return conn
def load_catalog(path: Path) -> list[CatalogEntry]:
    """Load the full product catalog into memory as ``CatalogEntry`` objects.

    Reads from ``data/etsy_orders.db`` (SQLite) when the database exists
    alongside the catalog file.  Falls back to parsing
    ``supplier_catalog.xlsx`` when the database has not yet been
    initialised — run ``python src/migrate_to_sqlite.py`` once to seed it.
    """
    db_path = path.parent / DB_FILE
    if db_path.exists():
        try:
            conn = get_db_connection(db_path)
            rows = conn.execute(
                "SELECT product_title, shop_name, stall, price, "
                "       charm_shop, charm_code, notes "
                "FROM   catalog ORDER BY id"
            ).fetchall()
            conn.close()
            entries = [
                CatalogEntry(
                    product_title = row["product_title"],
                    category      = "",               # not stored in SQLite schema
                    shop_name     = row["shop_name"]  or "",
                    stall         = row["stall"]       or "",
                    price         = row["price"]       or "",
                    charm_shop    = row["charm_shop"]  or "",
                    charm_code    = row["charm_code"]  or "",
                    notes         = row["notes"]       or "",
                )
                for row in rows
            ]
            log.info("Loaded %d products from SQLite catalog", len(entries))
            return entries
        except Exception as exc:
            log.warning(
                "SQLite catalog load failed (%s) — falling back to xlsx", exc
            )

    # ── xlsx fallback (used before migration or when DB is absent) ─────────
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[CATALOG_SHEET]
    h3 = str(ws.cell(1, 3).value or "").strip().lower()
    h7 = str(ws.cell(1, 7).value or "").strip().lower()
    has_category = h3 == "category"
    legacy_notes_first = has_category and h7 == "notes"

    entries: list[CatalogEntry] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[1]
        if not title or not isinstance(title, str):
            continue
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        if has_category:
            if legacy_notes_first:
                notes      = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                charm_shop = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                charm_code = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            else:
                charm_shop = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
                charm_code = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
                notes      = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            cat_v   = str(row[2]).strip() if row[2] is not None else ""
            shop_v  = str(row[3]).strip() if row[3] is not None else ""
            stall_v = str(row[4]).strip() if row[4] is not None else ""
            price_v = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        else:
            # Current 8-column layout: B title, C shop, D stall, E price, F/G charm, H notes
            cat_v      = ""
            shop_v     = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            stall_v    = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            price_v    = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            charm_shop = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
            charm_code = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
            notes      = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        entries.append(CatalogEntry(
            product_title = title.strip(),
            category      = cat_v,
            shop_name     = shop_v,
            stall         = stall_v,
            price         = price_v,
            notes         = notes,
            charm_shop    = charm_shop,
            charm_code    = charm_code,
        ))

    wb.close()
    log.info("Loaded %d products from catalog (xlsx fallback)", len(entries))
    return entries
PRODUCT_MAP_NUM_COLS = 8
_CHARM_COL_IDX   = 6
_CHARM_COL_LETTER = "F"
_CHARM_NAMED_RANGE = "CharmShopNames"
_CHARM_CODE_COL_IDX    = 7
_CHARM_CODE_COL_LETTER = "G"
_CHARM_CODES_NAMED_RANGE = "CharmCodes"
_SUPPLIER_SHOP_NAMED_RANGE  = "SupplierShopNames"
_SUPPLIER_STALL_NAMED_RANGE = "SupplierStalls"
_SUPPLIER_LIST_MAX_ROW      = 500
_CAT_CHARM_HDR_FILL = PatternFill("solid", fgColor="5B1A6B")   # same purple as route sheet
_CAT_CHARM_HDR_FONT = Font("Calibri", bold=True, color="FFFFFF", size=12)
_CAT_CHARM_PENDING_FILL = PatternFill("solid", fgColor="EFD9FC")
_CHARM_LIB_NOTE_HEIGHT = 188   # row height (points) for the instruction cell
def load_charm_library(path: Path) -> dict[str, CharmLibraryEntry]:
    """Return a ``charm_code → CharmLibraryEntry`` mapping (with photo BLOBs).

    Reads from ``data/etsy_orders.db`` (SQLite) when available; falls back
    to parsing the ``Charm Library`` sheet of ``supplier_catalog.xlsx``.
    """
    db_path = path.parent / DB_FILE
    if db_path.exists():
        try:
            conn = get_db_connection(db_path)
            rows = conn.execute(
                "SELECT code, sku, default_charm_shop, notes, photo "
                "FROM charm_library ORDER BY code"
            ).fetchall()
            conn.close()
            by_code: dict[str, CharmLibraryEntry] = {
                row["code"]: CharmLibraryEntry(
                    code               = row["code"],
                    sku                = row["sku"]                or "",
                    default_charm_shop = row["default_charm_shop"] or "",
                    notes              = row["notes"]              or "",
                    photo_bytes        = bytes(row["photo"]) if row["photo"] else None,
                )
                for row in rows
            }
            n_img = sum(1 for e in by_code.values() if e.photo_bytes)
            log.info(
                "Loaded %d charm(s) from SQLite (%d with photos)",
                len(by_code), n_img,
            )
            return by_code
        except Exception as exc:
            log.warning(
                "SQLite charm library load failed (%s) — falling back to xlsx", exc
            )

    # ── xlsx fallback ─────────────────────────────────────────────────────
    if not path.exists():
        return {}
    try:
        row_photos = extract_photos_from_xlsx(
            path, sheet_name=CHARM_LIBRARY_SHEET, photo_col_idx=0
        )
    except Exception as exc:
        log.warning("Charm Library photo extraction skipped: %s", exc)
        row_photos = {}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if CHARM_LIBRARY_SHEET not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[CHARM_LIBRARY_SHEET]
        by_code = {}
        for r_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not code or code.lower() == "charm code":
                continue
            sku_val  = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            def_shop = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            notes    = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            ent = CharmLibraryEntry(
                code=code,
                sku=sku_val,
                default_charm_shop=def_shop,
                notes=notes,
                photo_bytes=row_photos.get(r_num),
            )
            if code in by_code:
                log.warning(
                    "Duplicate Charm Code %r in %s — using last row wins",
                    code, CHARM_LIBRARY_SHEET,
                )
            by_code[code] = ent
        wb.close()
        n_img = sum(1 for e in by_code.values() if e.photo_bytes)
        log.info(
            "Loaded %d charm(s) from '%s' (%d with photos) (xlsx fallback)",
            len(by_code), CHARM_LIBRARY_SHEET, n_img,
        )
        return by_code
    except Exception as exc:
        log.warning("Could not load charm library from %s: %s", path.name, exc)
        return {}
def load_charm_shops(path: Path) -> list[CharmShop]:
    """Load the charm-shop reference list.

    Reads from ``data/etsy_orders.db`` (SQLite) when available; falls back
    to the ``Charm Shops`` sheet of ``supplier_catalog.xlsx``.
    Returns an empty list (with a warning) if neither source has data.
    """
    db_path = path.parent / DB_FILE
    if db_path.exists():
        try:
            conn = get_db_connection(db_path)
            rows = conn.execute(
                "SELECT shop_name, stall, notes FROM charm_shops ORDER BY id"
            ).fetchall()
            conn.close()
            shops = [
                CharmShop(
                    shop_name = row["shop_name"],
                    stall     = row["stall"] or "",
                    notes     = row["notes"] or "",
                )
                for row in rows
            ]
            log.info("Loaded %d charm shop(s) from SQLite", len(shops))
            return shops
        except Exception as exc:
            log.warning(
                "SQLite charm shops load failed (%s) — falling back to xlsx", exc
            )

    # ── xlsx fallback ─────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if CHARM_SHOPS_SHEET not in wb.sheetnames:
            wb.close()
            log.warning(
                "No '%s' sheet found in %s. "
                "Run with --init-charm-shops to create it.",
                CHARM_SHOPS_SHEET, path.name,
            )
            return []
        ws = wb[CHARM_SHOPS_SHEET]
        shops = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name  = str(row[0] or "").strip() if row[0]               else ""
            stall = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
            notes = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
            # Both name AND stall required; skips instructional/note rows.
            if name and stall:
                shops.append(CharmShop(shop_name=name, stall=stall, notes=notes))
        wb.close()
        log.info(
            "Loaded %d charm shop(s) from '%s' (xlsx fallback)",
            len(shops), CHARM_SHOPS_SHEET,
        )
        return shops
    except Exception as exc:
        log.warning("Could not load charm shops from %s: %s", path.name, exc)
        return []
def _needs_catalog_entry(r: ResolvedItem) -> bool:
    """
    Return True when an item is NOT reliably matched to a supplier location and
    should therefore be shown in the unmatched / awaiting-info section of the
    shopping route.

    This controls ROUTING only.  For catalog-completeness (whether the item
    needs its own row in supplier_catalog.xlsx) see _needs_own_catalog_row().

    Three cases qualify:

    1. Truly unmatched – no catalog entry found at all (r.supplier is None).

    2. False-positive against an EMPTY catalog entry (shop_name AND stall both
       empty, score < EMPTY_ENTRY_MATCH_THRESHOLD).
       Risk: item lands in "Awaiting Supplier Info" blue section but has no
       corresponding catalog row to fill in.  Adding its own row fixes the UX.

    3. False-positive against a FILLED catalog entry (shop_name or stall present,
       score < FILLED_ENTRY_MATCH_THRESHOLD).
       Risk: item inherits the WRONG supplier / stall and goes to the wrong shop.
       Threshold calibrated against the observed gap between the highest
       false-positive score (83.1) and the lowest genuine match score (93.7).
    """
    if r.supplier is None:
        return True
    # Authoritative supplier (operator-confirmed in the Unified Dashboard):
    # trust it unconditionally for routing — the fuzzy match-confidence checks
    # below only guard against false positives from automatic catalog matching,
    # which does not apply when the dashboard has explicitly resolved the shop.
    if getattr(r.supplier, "authoritative", False) and (
            r.supplier.shop_name or r.supplier.stall):
        return False
    # False positive against an empty-info catalog entry
    if (not r.supplier.shop_name and not r.supplier.stall
            and r.match_score < EMPTY_ENTRY_MATCH_THRESHOLD):
        return True
    # False positive against a filled catalog entry → wrong supplier risk
    if ((r.supplier.shop_name or r.supplier.stall)
            and r.match_score < FILLED_ENTRY_MATCH_THRESHOLD):
        return True
    return False
_SUPPLIERS_HEADER_ROW = (
    "ID",
    "Shop Name",
    "Mall",
    "Floor",
    "Stall",
    "Address",
    "Contact",
    "Notes",
)
@dataclass(frozen=True, slots=True)
class ProductMapPickerRow:
    """One Product Map row for UI pickers (photos + routing context)."""

    row_num: int
    title: str
    shop_name: str
    stall: str
    price: str = ""
    charm_shop: str = ""
    charm_code: str = ""
    notes: str = ""
def list_product_map_rows_for_picker(path: Path) -> list[ProductMapPickerRow]:
    """
    Return structured rows for each Product Map product (skips TOTAL).
    Used by the discontinued-product dialog (thumbnails + shop / stall).
    """
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[CATALOG_SHEET]
    except KeyError:
        wb.close()
        return []
    out: list[ProductMapPickerRow] = []
    for r_num in range(2, ws.max_row + 1):
        b_val = ws.cell(r_num, 2).value
        if not b_val or not isinstance(b_val, str):
            continue
        title = b_val.strip()
        if title.startswith("TOTAL:") or title == "Unknown Product":
            continue
        shop       = str(ws.cell(r_num, 3).value or "").strip()
        stall      = str(ws.cell(r_num, 4).value or "").strip()
        price      = str(ws.cell(r_num, 5).value or "").strip()
        charm_shop = str(ws.cell(r_num, 6).value or "").strip()
        charm_code = str(ws.cell(r_num, 7).value or "").strip()
        notes      = str(ws.cell(r_num, 8).value or "").strip()
        out.append(
            ProductMapPickerRow(
                row_num=r_num,
                title=title,
                shop_name=shop,
                stall=stall,
                price=price,
                charm_shop=charm_shop,
                charm_code=charm_code,
                notes=notes,
            )
        )
    wb.close()
    return out
def _build_catalog_photo_map(
    pairs: Iterable[tuple[str, bytes]],
) -> dict[str, bytes]:
    """Build the {normalized-title → photo} lookup used to apply canonical
    Product Map photos, with a collision-guarded 50-char prefix fallback.

    See :func:`get_catalog_photo_map` for the rationale.  Exact full-title
    keys are always registered.  A 50-char prefix key is registered only when
    exactly one distinct full title produces it AND it does not collide with an
    existing full-title key — otherwise the lookup would risk handing one
    product's photo to a different product that merely shares a title prefix.
    """
    result: dict[str, bytes] = {}
    short_titles: dict[str, set[str]] = defaultdict(set)
    short_photo:  dict[str, bytes] = {}
    for title, photo in pairs:
        key = _normalize(title)
        if not key:
            continue
        result[key] = photo                       # exact match — authoritative
        short = key[:50]
        short_titles[short].add(key)
        short_photo.setdefault(short, photo)
    for short, titles in short_titles.items():
        # Unambiguous prefix (one product) and not already a full-title key.
        if len(titles) == 1 and short not in result:
            result[short] = short_photo[short]
    return result
def get_catalog_photo_map(catalog_path: Path) -> dict[str, bytes]:
    """Return a mapping of *normalized product title* → canonical photo bytes.

    Used by the Orders Dashboard so every order for the same product always
    displays the same photo regardless of which PDF it was extracted from.

    Each product contributes:
      • The full normalized title (exact-match priority — always safe).
      • A 50-char prefix key (backward-compat for older cache entries) **only
        when that prefix unambiguously identifies a single product**.

    The 50-char prefix is intentionally collision-guarded: two genuinely
    different products can share the same first 50 normalized characters (e.g.
    "Kawaii Monchhichi MagSafe Case with Magnetic Grip Stand…" vs "…Grip &
    Beaded Charm…").  If such a prefix were exposed as a lookup key it would
    silently assign one product's photo to the other — the exact "wrong image
    on the order" class of bug.  We therefore drop any prefix shared by more
    than one distinct full title and fall back to the order's own (dashboard)
    image instead of guessing.

    Reads photo BLOBs from ``data/etsy_orders.db`` (SQLite) when the
    database exists; falls back to extracting them from column A of the
    Product Map sheet in ``supplier_catalog.xlsx``.
    """
    db_path = catalog_path.parent / DB_FILE
    if db_path.exists():
        try:
            conn = get_db_connection(db_path)
            rows = conn.execute(
                "SELECT product_title, photo FROM catalog WHERE photo IS NOT NULL"
            ).fetchall()
            conn.close()
            result = _build_catalog_photo_map(
                (row["product_title"], bytes(row["photo"])) for row in rows
            )
            log.info(
                "Catalog photo map: %d key(s) for %d product(s) with photos (SQLite)",
                len(result), len(rows),
            )
            return result
        except Exception as exc:
            log.warning(
                "SQLite catalog photo map failed (%s) — falling back to xlsx", exc
            )

    # ── xlsx fallback ─────────────────────────────────────────────────────
    if not catalog_path.exists():
        return {}
    try:
        rows = list_product_map_rows_for_picker(catalog_path)
        if not rows:
            return {}
        row_photos = extract_photos_from_xlsx(
            catalog_path, sheet_name=CATALOG_SHEET, photo_col_idx=0
        )
        pairs = [
            (row.title, row_photos[row.row_num])
            for row in rows
            if row_photos.get(row.row_num)
        ]
        result = _build_catalog_photo_map(pairs)
        log.info(
            "Catalog photo map: %d key(s) for %d product(s) with photos (xlsx fallback)",
            len(result), len(pairs),
        )
        return result
    except Exception as exc:
        log.warning("get_catalog_photo_map failed: %s", exc)
        return {}
def apply_canonical_charm_fields_to_resolved(
    resolved: list[ResolvedItem],
    catalog_path: Path,
) -> int:
    """Overwrite each item's ``r.supplier.charm_shop`` with the canonical shop
    from the Charm Library (``default_charm_shop``) for that charm code.

    After ``normalize_catalog_charm_shops`` runs, the Charm Library holds the
    single ground-truth shop for every charm code.  Cached items loaded from
    JSON may still carry the OLD per-order ``charm_shop`` value (from before
    normalisation), which makes the shopping-route charm aggregation split the
    same code into multiple rows — contradicting the dashboard.

    This function propagates the library's canonical shop back into every
    ``ResolvedItem`` so the Excel aggregation and the dashboard always agree.

    Returns the number of items whose ``charm_shop`` was corrected.
    """
    charm_lib = load_charm_library(catalog_path)
    if not charm_lib:
        return 0
    # Canonical shop per code (only codes that have one)
    canonical: dict[str, str] = {
        code: (entry.default_charm_shop or "").strip()
        for code, entry in charm_lib.items()
        if (entry.default_charm_shop or "").strip()
    }
    if not canonical:
        return 0
    corrected = 0
    for r in resolved:
        if not r.supplier:
            continue
        code = (r.supplier.charm_code or "").strip()
        if not code:
            continue
        target_shop = canonical.get(code)
        if target_shop and (r.supplier.charm_shop or "").strip() != target_shop:
            r.supplier.charm_shop = target_shop
            corrected += 1
    if corrected:
        log.info(
            "apply_canonical_charm_fields: corrected charm_shop on %d item(s) "
            "to match Charm Library canonical mapping",
            corrected,
        )
    return corrected
def apply_catalog_photos_to_resolved(
    resolved: list[ResolvedItem],
    catalog_path: Path,
    *,
    fill_missing_only: bool = False,
) -> int:
    """Apply the canonical Product Map photo to each item from the catalog.

    Two modes:

    * ``fill_missing_only=False`` (PDF flow — default): overwrite every item's
      ``photo_bytes`` with the catalog photo when one is available.  This
      normalises photos across PDF batches so the same product always looks the
      same regardless of which slip it was scanned from.

    * ``fill_missing_only=True`` (Unified Dashboard ``--import-json`` flow):
      treat the image supplied by the dashboard as the **single source of
      truth** and only *fill in* a catalog photo for items that arrived without
      one.  The dashboard already embeds, for every line item, the exact Etsy
      listing image it renders in its own order gallery (``image_b64``).
      Overwriting that with a title-matched catalog photo is what made the
      route Excel disagree with the dashboard (e.g. order #4070998188 showed the
      curated Product Map photo instead of the live listing image the operator
      sees).  In this mode the route and the dashboard are guaranteed to display
      identical per-order images.

    Returns the number of items whose photo was updated.
    """
    catalog_photos = get_catalog_photo_map(catalog_path)
    if not catalog_photos:
        return 0
    updated = 0
    skipped = 0
    for r in resolved:
        # Dashboard-supplied images are authoritative — never replace them.
        if fill_missing_only and r.item.photo_bytes:
            skipped += 1
            continue
        norm = _normalize(r.item.title)
        canonical = catalog_photos.get(norm) or catalog_photos.get(norm[:50])
        if canonical and canonical != r.item.photo_bytes:
            r.item.photo_bytes = canonical
            updated += 1
    if fill_missing_only:
        log.info(
            "apply_catalog_photos: kept %d dashboard-supplied image(s) as-is; "
            "filled %d missing photo(s) from the Product Map",
            skipped, updated,
        )
    elif updated:
        log.info(
            "apply_catalog_photos: replaced photos for %d of %d item(s) "
            "with canonical Product Map images",
            updated, len(resolved),
        )
    return updated
_DISC_HDR_FILL = PatternFill("solid", fgColor="7F1D1D")
_DISC_HDR_FONT = Font("Calibri", bold=True, color="FFFFFF", size=11)
_DISC_BODY     = Font("Calibri", size=10)
_DISC_BODY_BOLD = Font("Calibri", bold=True, size=10)
_DISC_ROW_FILL = PatternFill("solid", fgColor="FEF2F2")
_DISC_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DISC_WRAP     = Alignment(vertical="center", wrap_text=True)
_DISC_HEADERS = [
    "PHOTO", "PRODUCT TITLE", "SHOP NAME", "STALL",
    "PRICE", "CHARM SHOP", "CHARM CODE", "NOTES",
    "DISCONTINUED DATE",
]
_DISC_COL_WIDTHS = {
    "A": PHOTO_COL_W,  # match Product Map — wide enough for PHOTO_PX embeds
    "B": 48.0, "C": 14.0, "D": 14.0, "E": 10.0,
    "F": 9.0, "G": 14.0, "H": 26.0, "I": 20.0,
}
def extract_photos_from_xlsx(
    xlsx_path: Path,
    sheet_name: str = "Shopping Route",
    photo_col_idx: int = 1,     # 0-based column index of the Photo column (B = 1)
) -> dict[int, bytes]:
    """
    Extract embedded images from *sheet_name* in an xlsx file.

    Returns {excel_row_1based: jpeg_bytes} for every image anchored to
    *photo_col_idx*.  Works with both oneCellAnchor and twoCellAnchor.

    Because .xlsx files are ZIP archives we can read images directly without
    relying on openpyxl's (limited) image-read support.
    """
    _R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    _XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    _A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    def _zip_path(target: str, base_dir: str = 'xl') -> str:
        """
        Normalise a relationship Target value to a bare zip entry path.
        Handles absolute (/xl/...), relative (../...), and plain names.
        """
        t = target.strip()
        if t.startswith('/'):
            return t[1:]                         # absolute: strip leading /
        if t.startswith('../'):
            parent = base_dir.rsplit('/', 1)[0]  # go up one dir
            return parent + '/' + t[3:]
        if '/' not in t:
            return base_dir + '/' + t            # bare filename in same dir
        if not t.startswith('xl/'):
            return 'xl/' + t
        return t

    result: dict[int, bytes] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            all_files = set(zf.namelist())

            # ---- Step 1: find the worksheet rId for sheet_name ----
            wb_tree  = ET.parse(zf.open('xl/workbook.xml'))
            sheet_rId: str | None = None
            for elem in wb_tree.getroot().iter():
                if elem.tag.endswith('}sheet') or elem.tag == 'sheet':
                    if elem.get('name') == sheet_name:
                        sheet_rId = elem.get(f'{{{_R}}}id') or elem.get('r:id')
                        break
            if not sheet_rId:
                return result

            # ---- Step 2: rId → worksheet XML zip path ----
            wb_rels = 'xl/_rels/workbook.xml.rels'
            if wb_rels not in all_files:
                return result
            sheet_file: str | None = None
            for rel in ET.parse(zf.open(wb_rels)).getroot():
                if rel.get('Id') == sheet_rId:
                    sheet_file = _zip_path(rel.get('Target', ''), 'xl')
                    break
            if not sheet_file or sheet_file not in all_files:
                return result

            # ---- Step 3: worksheet rels → drawing XML zip path ----
            ws_dir  = sheet_file.rsplit('/', 1)[0]   # e.g. 'xl/worksheets'
            ws_name = sheet_file.rsplit('/', 1)[1]
            ws_rels = f'{ws_dir}/_rels/{ws_name}.rels'
            if ws_rels not in all_files:
                return result
            drawing_file: str | None = None
            for rel in ET.parse(zf.open(ws_rels)).getroot():
                if 'drawing' in rel.get('Type', '').lower():
                    drawing_file = _zip_path(rel.get('Target', ''), ws_dir)
                    break
            if not drawing_file or drawing_file not in all_files:
                return result

            # ---- Step 4: drawing rels → rId to image zip path ----
            dr_dir  = drawing_file.rsplit('/', 1)[0]  # e.g. 'xl/drawings'
            dr_name = drawing_file.rsplit('/', 1)[1]
            dr_rels = f'{dr_dir}/_rels/{dr_name}.rels'
            if dr_rels not in all_files:
                return result
            rId_to_img: dict[str, str] = {}
            for rel in ET.parse(zf.open(dr_rels)).getroot():
                rid      = rel.get('Id', '')
                img_path = _zip_path(rel.get('Target', ''), dr_dir)
                rId_to_img[rid] = img_path

            # ---- Step 5: parse drawing anchors → row → bytes ----
            d_root = ET.parse(zf.open(drawing_file)).getroot()
            for anchor in d_root:
                a_tag = anchor.tag.split('}')[-1]
                if a_tag not in ('oneCellAnchor', 'twoCellAnchor'):
                    continue
                fr = anchor.find(f'{{{_XDR}}}from')
                if fr is None:
                    continue
                col_e = fr.find(f'{{{_XDR}}}col')
                row_e = fr.find(f'{{{_XDR}}}row')
                if col_e is None or row_e is None:
                    continue
                if int(col_e.text) != photo_col_idx:
                    continue
                excel_row = int(row_e.text) + 1   # 0-based → 1-based

                blip = anchor.find(f'.//{{{_A}}}blip')
                if blip is None:
                    continue
                r_embed   = blip.get(f'{{{_R}}}embed', '')
                img_path  = rId_to_img.get(r_embed, '')
                if img_path and img_path in all_files:
                    result[excel_row] = zf.read(img_path)

    except Exception as e:
        log.warning("extract_photos_from_xlsx(%s): %s", xlsx_path.name, e)

    if result:
        log.info("Extracted %d embedded photo(s) from %s", len(result), xlsx_path.name)
    return result
def load_existing_statuses(xlsx_path: Path) -> dict[tuple[str, str], str]:
    """
    Read the Shopping Route sheet from an existing output file and return every
    non-default component status so they can be re-applied after a re-generate.

    Returns {(order_number, component): status_string}
    where component is one of 'case', 'grip', 'charm'.
    Only 'Purchased', 'Out of Stock', and 'Out of Production' are preserved
    (Pending and N/A are defaults that will be restored automatically).
    """
    if not xlsx_path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Shopping Route" not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb["Shopping Route"]
    except Exception as e:
        log.warning("Could not read existing statuses from %s: %s", xlsx_path.name, e)
        return {}

    keep = {"Purchased", "Out of Stock", "Out of Production"}
    _items_vals = {"case only", "grip only", "case, grip", "\u2014",
                  "仅手机壳", "仅支架", "手机壳、支架"}
    statuses: dict[tuple[str, str], str] = {}

    for row_vals in ws.iter_rows(min_row=5, values_only=True):
        if not row_vals or len(row_vals) < 12:
            continue
        # Detect format: new has "Items to Purchase" col 7 (index 6)
        is_new_format = (
            len(row_vals) >= 14
            or (len(row_vals) >= 13 and str(row_vals[6] or "").strip() in _items_vals)
        )
        if is_new_format:
            order_cell = row_vals[12] if len(row_vals) > 12 else None
            case_col, grip_col, charm_col = 7, 8, 9
        else:
            order_cell = row_vals[11] if len(row_vals) > 11 else None
            case_col, grip_col, charm_col = 6, 7, 8
        if not order_cell or not isinstance(order_cell, str):
            continue

        order_str = order_cell.strip()

        # col 6 = Product title — used to distinguish items within the same order
        title_val  = str(row_vals[5]).strip() if row_vals[5] else ""
        norm_title = _normalize(title_val)[:50]

        # Aggregated charm row: ~C:<charm_code>
        if order_str.startswith("~C:"):
            charm_code_key = order_str[3:].strip()
            if charm_code_key and len(row_vals) > charm_col:
                val = str(row_vals[charm_col]).strip() if row_vals[charm_col] else ""
                if val in keep:
                    statuses[(charm_code_key, "", "charm_agg")] = val
            continue

        # Legacy per-order charm row: ~#<order_num> (backward compat)
        if order_str.startswith("~#") or order_str.startswith("~?#"):
            prefix_len = 3 if order_str.startswith("~?#") else 2
            charm_order_num = order_str[prefix_len:].strip()
            if charm_order_num.isdigit() and len(row_vals) > charm_col:
                val = str(row_vals[charm_col]).strip() if row_vals[charm_col] else ""
                if val in keep:
                    statuses[(charm_order_num, norm_title, "charm")] = val
            continue

        # Regular case/grip supplier row: "#XXXXXXX"
        order_num = order_str.lstrip("#").strip()
        if not order_num.isdigit():
            continue
        for comp, col_idx in (("case", case_col), ("grip", grip_col)):
            val = str(row_vals[col_idx]).strip() if row_vals[col_idx] else ""
            if val in keep:
                statuses[(order_num, norm_title, comp)] = val

    wb.close()
    if statuses:
        log.info("Preserved %d non-default status value(s) from existing file",
                 len(statuses))
    return statuses
def _load_ui_status_cache(
    output_dir: Path,
) -> dict[tuple[str, str, str], str]:
    """Load purchase-status overrides written by the Orders Dashboard UI.

    The UI saves statuses to ``<output_dir>/route_statuses_cache.json`` with
    keys serialised as ``order_num\\x00norm_title\\x00comp``.  This function
    reads that file and returns the same dict format as
    ``load_existing_statuses`` so the two can be merged transparently.

    Returns an empty dict silently if the file does not exist or cannot be
    parsed — callers should treat it as a best-effort supplement.
    """
    cache_path = output_dir / "route_statuses_cache.json"
    if not cache_path.exists():
        return {}
    try:
        import json as _json
        raw: dict = _json.loads(cache_path.read_text(encoding="utf-8"))
        # "Pending" is included so an explicit UI reset to Pending can override
        # a stale "Purchased" value that load_existing_statuses may have read
        # from the previous shopping_route.xlsx.  Without "Pending" here, the
        # xlsx stale value would win and the item would appear fully-purchased.
        valid = {"Pending", "Purchased", "Out of Stock", "Out of Production"}
        result: dict[tuple[str, str, str], str] = {}
        for k_str, val in raw.items():
            if val not in valid:
                continue
            parts = k_str.split("\x00", 2)
            if len(parts) != 3:
                continue
            order_num, norm_title, comp = parts
            # Enforce the same [:50] truncation used throughout the sheet
            # generation pipeline so keys align with the xlsx-based statuses.
            result[(order_num, norm_title[:50], comp)] = val
        if result:
            log.info(
                "UI status cache: loaded %d override(s) from %s",
                len(result), cache_path.name,
            )
        return result
    except Exception as e:
        log.warning("Could not read UI status cache %s: %s", cache_path.name, e)
        return {}
# ══ STALL LOCATION ══
# Mirror of src/route/stall-location.js. We shop several markets, and a stall
# outside the home one carries its market as a prefix ("康乐北区5A40-42").
# Splitting that off is what lets the route be walked one BUILDING at a time
# rather than bouncing between markets that share a floor number, and it is why
# the floor is read from the code AFTER the prefix is removed.
_HOME_BUILDING_ID = "tongxin"

# (id, walking order, aliases as typed into the catalog). Longest alias wins, so
# "康乐北区" can never be swallowed by "康乐".
_BUILDINGS: list[tuple[str, int, tuple[str, ...]]] = [
    ("jingji",       10, ("经济",)),
    ("tongxin",      20, ("通信", "通信城", "通信市场")),
    ("kangle-north", 30, ("康乐北区", "康乐北")),
    # 康乐 is the same complex as 康乐北区, so it stays pinned beside it.
    ("kangle",       31, ("康乐",)),
    ("taipingyang",  40, ("太平洋",)),
    ("huitong",      50, ("汇通",)),
    # The charm market. Bare codes, no prefix to parse — only Shopping Mode
    # re-homes charms to it, so it carries no alias here.
    ("longsheng",    60, ()),
]
_UNREGISTERED_ORDER = 9000
_UNLOCATED_ORDER = 9999
_UNKNOWN_FLOOR = 999
_PLACEHOLDER_CODES = {"-", "???"}
_ALIAS_INDEX = sorted(
    ((alias, bid, order) for bid, order, aliases in _BUILDINGS for alias in aliases),
    key=lambda a: -len(a[0]),
)


def _normalize_stall(stall: str) -> str:
    """Fold full-width forms and exotic dashes so one spelling reaches the rules."""
    return re.sub(r"[–—―]", "-", unicodedata.normalize("NFKC", str(stall or ""))).strip()


def _floor_from_code(code: str) -> int:
    """Floor of a stall code whose building prefix has already been removed."""
    if not code:
        return _UNKNOWN_FLOOR
    if re.match(r"^A2", code, re.IGNORECASE):
        return 2
    m = re.match(r"^(\d)", code)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d)[A-Za-z]", code)
    if m:
        return int(m.group(1))
    return _UNKNOWN_FLOOR


def _parse_stall(stall: str) -> tuple[str, int, str, int]:
    """
    Split a stall code into (building id, walking order, local code, floor).

    A registered prefix wins; failing that, a leading run of non-ASCII is an
    unregistered market and becomes a location of its own, so a newly typed
    market never silently files itself under the home one.
    """
    raw = _normalize_stall(stall)
    if not raw or raw in _PLACEHOLDER_CODES:
        return "", _UNLOCATED_ORDER, "", _UNKNOWN_FLOOR
    for alias, bid, order in _ALIAS_INDEX:
        if raw[: len(alias)].upper() == alias.upper():
            code = raw[len(alias):].strip()
            return bid, order, code, _floor_from_code(code)
    m = re.match(r"^([^\x00-\x7F]+)", raw)
    if m:
        code = raw[len(m.group(1)):].strip()
        return "x:" + m.group(1).strip(), _UNREGISTERED_ORDER, code, _floor_from_code(code)
    home_order = next(order for bid, order, _ in _BUILDINGS if bid == _HOME_BUILDING_ID)
    return _HOME_BUILDING_ID, home_order, raw, _floor_from_code(raw)


def _stall_floor(stall: str) -> int:
    """
    Parse the floor number from a stall code for ascending-floor sort order.

    Conventions observed in the catalog:
      A2xxx / A2-xx  -> 2nd floor (A-block)
      4Cxx  / 4Dxx   -> 4th floor
      5Xxx  / 5Cxx   -> 5th floor
      A market prefix ("康乐北区4D32") is stripped before these rules apply.

    Returns 999 for unknown stalls so they sort to the very end.
    """
    return _parse_stall(stall)[3]


def _stall_sort_key(stall: str, shop: str) -> tuple:
    """
    The order the route is walked: market, then floor, then stall, then shop.
    Shared by every sheet so the workbook, the desktop table and Shopping Mode
    all send the shopper round in the same sequence.
    """
    bid, order, code, floor = _parse_stall(stall)
    return (
        order,
        bid,
        floor if floor != _UNKNOWN_FLOOR else 9999,
        code.lower() or "\uffff",
        (shop or "").lower(),
    )
# ══ END STALL LOCATION ══
def _style_has(style: str) -> tuple[bool, bool, bool]:
    """Return (has_case, has_grip, has_charm) booleans from a style string."""
    s = style.lower()
    # "stand" / "kickstand" (e.g. "Case+Stand+Charm", "Kickstand Cover") = grip
    has_grip = "grip" in s or "stand" in s
    return "case" in s, has_grip, "charm" in s
def _style_flags(style: str) -> tuple[str, str, str]:
    """
    Return checkmark strings for the Case/Grip/Charm columns (Orders Detail).
    Empty string means the component is not included.
    """
    has_case, has_grip, has_charm = _style_has(style)
    return (
        "\u2713" if has_case  else "",
        "\u2713" if has_grip  else "",
        "\u2713" if has_charm else "",
    )
def _section_complete(status: str | None) -> bool:
    """
    True if the component status indicates procurement is complete.
    Both "Purchased" and "Out of Production" are terminal — no further action.
    """
    return status in ("Purchased", "Out of Production")
# ── Unified-Dashboard round-trip key ──────────────────────────────────────────
# The Chinese status workbook (shopping_route_zh_status.xlsx) is grouped by
# supplier/floor and carries NO order number or product title in any visible
# cell, so an employee's edited file cannot be mapped back to a specific order
# line by content alone. To make the file losslessly re-importable, we embed a
# hidden machine-readable key column whose header is _UED_KEY_HEADER. Each data
# row stores a JSON object identifying exactly which dashboard line (or charm
# aggregate) it represents, so POST /api/route/import-status can write the edited
# statuses straight back into route_assignments with zero ambiguity. The column
# is hidden + excluded from the auto-filter/print area, so the sheet looks
# identical to before for the human reading it.
_UED_KEY_HEADER = "__UED_KEY__"


def _ued_line_item_key(title: str, listing_id: str) -> str:
    """Reproduce the dashboard's ``lineItemKey`` (src/route/dashboard.js).

    Format: ``<normalized title [:50]>#L<listing_id>`` — falling back to the
    bare 50-char title key when no listing id is present. MUST stay byte-for-byte
    identical to the Node implementation so the imported statuses land on the
    same route_assignments rows the dashboard created.
    """
    base = _normalize(title)[:50]
    lid = str(listing_id or "").strip()
    return f"{base}#L{lid}" if lid else base


def _charm_status_key(r: ResolvedItem) -> tuple[str, str, str]:
    """Same key shape as load_existing_statuses / route_statuses_cache (50-char norm)."""
    return (
        str(r.order.order_number).strip(),
        _normalize(r.item.title)[:50],
        "charm",
    )
def _charm_line_in_shopping_route(
    r: ResolvedItem,
    statuses: dict[tuple[str, str, str], str] | None,
) -> bool:
    """True when this line item should appear in the route Charm section (Excel/HTML).

    Without this filter, every order line with a charm in ``style`` was included,
    so Purchased / Out-of-Production charms still inflated counts and shop lists
    even though the dashboard already showed them as done.
    """
    if not _style_has(r.item.style)[2]:
        return False
    st = (statuses or {}).get(_charm_status_key(r), "Pending")
    return not _section_complete(st)
def _load_dashboard_route_exclusions(path: Path) -> set[tuple[str, str]]:
    """Parse ``--exclude-orders-file`` JSON from the Orders Dashboard.

    Returns a set of ``(normalized_title, order_number_str)`` for fast lookup.
    """
    import json as _json
    raw = _json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        log.warning(
            "exclude-orders-file: expected a JSON array, got %s — ignoring",
            type(raw).__name__,
        )
        return set()
    out: set[tuple[str, str]] = set()
    for e in raw:
        if not isinstance(e, dict):
            continue
        o = str(e.get("order", "")).strip()
        nt = _normalize(str(e.get("norm_title", "")))
        if o and nt:
            out.add((nt, o))
    return out
def _items_to_purchase(
    has_case: bool,
    has_grip: bool,
    case_status: str | None,
    grip_status: str | None,
    lang: str = "en",
) -> str:
    """
    Return a concise label for the "Items to Purchase" column.

    Components needing purchase: Pending, Out of Stock, or Out of Production.
    Returns: "case only", "grip only", "case, grip", or "—" when nothing needed.
    """
    needs_action = {"Pending", "Out of Stock", "Out of Production"}
    case_needs = has_case and (case_status or "Pending") in needs_action
    grip_needs = has_grip and (grip_status or "Pending") in needs_action
    if case_needs and grip_needs:
        return _t("case, grip", lang) if lang == "zh" else "case, grip"
    if case_needs:
        return _t("case only", lang) if lang == "zh" else "case only"
    if grip_needs:
        return _t("grip only", lang) if lang == "zh" else "grip only"
    return "\u2014"  # em dash
def _normalize(text: str) -> str:
    text = text.replace("|", ",")
    return re.sub(r"\s+", " ", text).strip().lower()
def _route_item_sort_key(r: ResolvedItem) -> tuple[str, str, str]:
    """Within one supplier stop: stack rows with the same product title together.

    Order is independent of Case/Grip/Charm status so \"Out of Stock\" and
    \"Pending\" lines for the same product appear consecutively.
    """
    return (
        _normalize(r.item.title),
        r.item.phone_model or "",
        r.order.order_number,
    )
def get_fts_candidates(
    conn: sqlite3.Connection,
    normalized_title: str,
    limit: int = 15,
) -> list[str]:
    """Return up to *limit* ``product_title`` strings from the FTS5 trigram index.

    Results are ordered by BM25 relevance (most similar first).  The caller
    should apply ``_normalize`` to each returned title before fuzzy scoring.

    Strategy: each word in the query is phrase-quoted individually so that FTS5
    operator keywords (AND, OR, NOT, ``*``, ``-``) cannot be mis-interpreted.
    Words shorter than 3 characters are skipped because the trigram tokenizer
    cannot generate a trigram from them.  We first try an AND query (all words
    must appear) for precision; if that returns nothing we fall back to an OR
    query (any word appears) for better recall.  The final rapidfuzz pass then
    re-ranks/filters the small candidate set by score threshold.

    Returns an empty list on any error so callers can fall back gracefully.
    """
    import re as _re

    # Extract clean alphanumeric words long enough for trigram matching.
    words = [_re.sub(r"[^a-z0-9]", "", w) for w in normalized_title.lower().split()]
    words = [w for w in words if len(w) >= 3]
    if not words:
        return []

    _SQL = """
        SELECT c.product_title
        FROM   catalog_fts f
        JOIN   catalog     c ON f.rowid = c.id
        WHERE  catalog_fts MATCH ?
        ORDER  BY bm25(catalog_fts)
        LIMIT  ?
    """

    def _run(join_op: str) -> list[str]:
        quoted = ['"{}"'.format(w.replace('"', '""')) for w in words]
        safe_term = f" {join_op} ".join(quoted)
        try:
            rows = conn.execute(_SQL, (safe_term, limit)).fetchall()
            return [row["product_title"] for row in rows]
        except Exception as exc:
            log.debug("FTS5 query (%s) failed for %r: %s", join_op, normalized_title, exc)
            return []

    results = _run("AND")
    if not results:
        results = _run("OR")
    return results
def match_items(
    orders: list[Order],
    catalog: list[CatalogEntry],
    threshold: int,
    *,
    db_path: Path | None = None,
) -> list[ResolvedItem]:
    """Match every order line-item against the product catalog.

    **FTS5 path** (when *db_path* points to a live SQLite database):

    1. ``get_fts_candidates()`` fires a single indexed query against the
       ``catalog_fts`` trigram table and returns ≤15 candidate titles in
       BM25 order — O(1) lookup regardless of catalog size.
    2. ``rapidfuzz.process.extractOne`` scores only those 15 candidates
       instead of the full catalog, keeping all existing thresholds
       (``EMPTY_ENTRY_MATCH_THRESHOLD``, ``FILLED_ENTRY_MATCH_THRESHOLD``,
       ``SAME_PRODUCT_THRESHOLD``, ``VARIANT_HINT_THRESHOLD``) intact.

    **O(N) fallback** (when the database is absent, unavailable, or when
    FTS returns no candidates for a particular title):

    The original full-catalog ``process.extractOne`` scan is used so the
    result is identical to the pre-migration behaviour.
    """
    catalog_titles = [_normalize(e.product_title) for e in catalog]

    # Pre-build a normalized-title → catalog-index map for O(1) FTS lookups.
    # When duplicate normalized titles exist the last index wins — consistent
    # with extractOne picking the last match on ties.
    title_to_idx: dict[str, int] = {t: i for i, t in enumerate(catalog_titles)}

    # Open one DB connection for the whole batch; None means O(N) fallback.
    conn: sqlite3.Connection | None = None
    if db_path is not None and db_path.exists():
        try:
            conn = get_db_connection(db_path)
        except Exception as exc:
            log.warning(
                "match_items: cannot open SQLite DB (%s) — using O(N) scan", exc
            )

    fts_active = conn is not None
    resolved: list[ResolvedItem] = []

    for order in orders:
        for item in order.items:
            norm = _normalize(item.title)

            # ── FTS5 pre-filter path ───────────────────────────────────────
            if conn is not None:
                candidate_titles = get_fts_candidates(conn, norm)
                if candidate_titles:
                    # Map each FTS candidate title back to its in-memory
                    # CatalogEntry.  Products added mid-session via
                    # update_catalog() live in `catalog` but may not yet be
                    # in the FTS index; they are handled by the fallback below.
                    cand_norm: list[str] = []
                    cand_entries: list[CatalogEntry] = []
                    for raw_title in candidate_titles:
                        n = _normalize(raw_title)
                        idx = title_to_idx.get(n)
                        if idx is not None:
                            cand_norm.append(n)
                            cand_entries.append(catalog[idx])

                    if cand_norm:
                        result = process.extractOne(
                            norm,
                            cand_norm,
                            scorer=fuzz.token_sort_ratio,
                            score_cutoff=threshold,
                        )
                        if result:
                            _, score, ci = result
                            resolved.append(
                                ResolvedItem(order, item, cand_entries[ci], score)
                            )
                        else:
                            resolved.append(ResolvedItem(order, item, None, 0.0))
                        continue   # item handled — skip the O(N) scan below

            # ── O(N) fallback: full catalog scan ──────────────────────────
            # Used when: DB absent | FTS returned no candidates | candidate
            # titles did not map to any in-memory entry.
            result = process.extractOne(
                norm,
                catalog_titles,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=threshold,
            )
            if result:
                _, score, idx = result
                resolved.append(ResolvedItem(order, item, catalog[idx], score))
            else:
                resolved.append(ResolvedItem(order, item, None, 0.0))

    if conn is not None:
        conn.close()

    matched = sum(1 for r in resolved if r.supplier)
    log.info(
        "Matched %d / %d items (threshold %d%%, %s)",
        matched, len(resolved), threshold,
        "FTS5+rapidfuzz" if fts_active else "O(N) rapidfuzz",
    )
    return resolved
_CHARM_IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp")
def _disk_charm_files_index(root: Path) -> dict[str, tuple[str, str]]:
    """
    Map charm code (filename stem) → (path relative to *root* with /, basename).
    One file per stem; extension priority left-to-right in _CHARM_IMAGE_EXTENSIONS.
    """
    best: dict[str, tuple[str, str]] = {}
    for ext in _CHARM_IMAGE_EXTENSIONS:
        for path in root.rglob(f"*{ext}"):
            if not path.is_file():
                continue
            st = path.stem
            if not st or st.lower() in ("charm code", "photo"):
                continue
            if st in best:
                continue
            best[st] = (path.relative_to(root).as_posix(), path.name)
    return best
def _atomic_write_text(path: Path, text: str, *, encoding: str = "utf-8") -> None:
    """Write *text* to *path* via a same-directory ``*.tmp`` rename (crash-safe)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    try:
        tmp.write_text(text, encoding=encoding)
        tmp.replace(path)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise
def export_charm_manifest(
    catalog_path: Path,
    charm_images_dir: Path | None,
    output_path: Path,
    *,
    route_snapshot: dict | None = None,
) -> int:
    """
    Emit JSON merging Charm Library rows and files under *charm_images_dir*
    (including subfolders).  Intended for websites, imports, or audits.

    Writes atomically so consumers never read a half-written JSON file.
    """
    charms: dict[str, dict] = {}

    if catalog_path.exists():
        try:
            wb = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)
            if CHARM_LIBRARY_SHEET in wb.sheetnames:
                ws = wb[CHARM_LIBRARY_SHEET]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    if not code or code.lower() == "charm code":
                        continue
                    charms[code] = {
                        "code": code,
                        "sku": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                        "default_charm_shop": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                        "notes": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                        "sources": ["library"],
                    }
            wb.close()
        except Exception as exc:
            log.warning("Manifest: catalog read failed: %s", exc)

    img_root = charm_images_dir
    if img_root and img_root.is_dir():
        for st, (rel, fname) in _disk_charm_files_index(img_root).items():
            rec = charms.get(st)
            if rec:
                if "disk" not in rec["sources"]:
                    rec["sources"].append("disk")
                rec["image_file"] = fname
                rec["image_relative"] = rel
            else:
                charms[st] = {
                    "code": st,
                    "sku": "",
                    "default_charm_shop": "",
                    "notes": "",
                    "sources": ["disk"],
                    "image_file": fname,
                    "image_relative": rel,
                }

    manifest: dict = {
        "version": 1,
        "schema": "charm_manifest",
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "charm_images_dir": str(img_root.resolve()) if img_root else None,
        "supplier_catalog": str(catalog_path.resolve()) if catalog_path.exists() else None,
        "convention": (
            "Stable codes: PREFIX + zero-padded digits (e.g. CH-00001). "
            "New sequences use at least five digits when no prior PREFIX+digits codes exist; "
            "otherwise padding matches existing rows/files and widens past 9999 as needed. "
            "Filename = <code>.png|.jpg|.jpeg|.webp; subfolders allowed. "
            "Charm Library column C (SKU) holds the short label or stock code."
        ),
        "charm_codes_range_rows": CHARM_CODES_LIST_MAX_ROW,
        "charms": sorted(charms.values(), key=lambda x: x["code"].lower()),
    }
    if route_snapshot:
        manifest["route_snapshot"] = route_snapshot
    _atomic_write_text(
        output_path,
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log.info("Wrote charm manifest  ->  %s  (%d charm(s))", output_path, len(charms))
    return len(charms)
def _sanitize_charm_code_for_filename(code: str) -> str:
    """Filename stem for disk assets; strip and replace Windows-forbidden characters."""
    s = code.strip()
    for ch in '<>:"/\\|?*':
        s = s.replace(ch, "_")
    return s
def charm_photo_bytes_from_folder(code: str, root: Path | None) -> bytes | None:
    """
    Load ``<root>/<Charm Code>.<ext>`` (png/jpg/jpeg/webp), then any same-named
    file under *root* subfolders (``rglob``).  Returns None if not found.
    """
    if not root or not root.is_dir():
        return None
    stem = _sanitize_charm_code_for_filename(code)
    if not stem:
        return None

    def _read(p: Path) -> bytes | None:
        try:
            return p.read_bytes()
        except OSError as exc:
            log.warning("Could not read charm image %s: %s", p, exc)
            return None

    for ext in _CHARM_IMAGE_EXTENSIONS:
        direct = root / f"{stem}{ext}"
        if direct.is_file():
            return _read(direct)

    for ext in _CHARM_IMAGE_EXTENSIONS:
        matches = sorted(root.rglob(f"{stem}{ext}"))
        if len(matches) > 1:
            log.warning(
                "Multiple files for charm code %r — using %s",
                code, matches[0],
            )
        if matches:
            data = _read(matches[0])
            if data:
                return data
    return None
def _charm_photo_for_code(
    code: str,
    charm_library: dict[str, CharmLibraryEntry] | None,
    charm_images_dir: Path | None = None,
) -> bytes | None:
    """Resolve a charm's photo by code using the SAME source of truth as the
    Orders Sorting Dashboard, so the route Excel/HTML and the dashboard gallery
    can never show different images for the same charm.

    Source-of-truth ordering (critical — do not reorder):

      1. On-disk ``data/charm_images/<code>.<ext>`` — this is exactly what the
         dashboard reads via ``GET /api/route/charm-image`` AND exactly what it
         (re)writes whenever the operator uploads/replaces a charm image
         (``saveCharmImage`` → ``<code>.<ext>``).  It is therefore the single
         live source of truth and MUST win.

      2. Charm Library BLOB in the SQLite mirror (``charm_library.photo``) — a
         fallback only.  This BLOB is seeded once and is NOT refreshed when the
         operator replaces an image from the dashboard, so it goes stale; it may
         only be used for legacy codes that have no on-disk file yet.

    Previously the BLOB was consulted first, which is why a charm updated in the
    dashboard kept showing its OLD image in the generated Excel.
    """
    code = (code or "").strip()
    if not code:
        return None
    disk = charm_photo_bytes_from_folder(code, charm_images_dir)
    if disk:
        return disk
    if charm_library:
        ent = charm_library.get(code)
        if ent and ent.photo_bytes:
            return ent.photo_bytes
    return None
_THIN   = Side(style="thin", color="C0C0C0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT   = Font("Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font("Calibri", bold=True, size=16, color="1F4E79")
_SUB_FONT   = Font("Calibri", size=11, color="555555", italic=True)
_BODY       = Font("Calibri", size=10)
_BODY_BOLD  = Font("Calibri", bold=True, size=10)
_WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")
_WARN_FONT  = Font("Calibri", size=10, color="856404")
_SEC_FONT   = Font("Calibri", bold=True, size=12, color="1F4E79")
_GROUP_FILLS = [
    PatternFill("solid", fgColor="EBF2FA"),
    PatternFill("solid", fgColor="FFFFFF"),
]
_NA_FILL = PatternFill("solid", fgColor="EFEFEF")
_NA_FONT = Font("Calibri", size=9, color="AAAAAA", italic=True)
_NEEDSINFO_FILL = PatternFill("solid", fgColor="D9EAF7")   # pale blue
_NEEDSINFO_FONT = Font("Calibri", size=10, color="1F4E79")  # dark navy
_CHARM_BANNER_FILL  = PatternFill("solid", fgColor="3D1359")  # deep purple  – section banner
_CHARM_SHOPS_FILL   = PatternFill("solid", fgColor="EFD9FC")  # light lavender – shop-list row
_CHARM_HDR_FILL     = PatternFill("solid", fgColor="5B1A6B")  # mid purple   – sub-header
_CHARM_GROUP_FILLS  = [
    PatternFill("solid", fgColor="F8F0FD"),   # very light lavender (odd rows)
    PatternFill("solid", fgColor="FFFFFF"),   # white                (even rows)
]
# Per-charm-shop row tints: every charm that ships from the SAME supplier gets
# the SAME background colour, so a shopper can see at a glance which rows belong
# to one stall. A palette of soft, readable pastels is cycled in first-seen
# order; unassigned charms (no shop) stay neutral white. These are base fills —
# the Purchased / Out-of-Stock / Out-of-Production status colours are applied on
# top via conditional formatting and still win for already-handled rows.
_CHARM_SHOP_FILLS = [
    PatternFill("solid", fgColor="EDE3FB"),   # lavender
    PatternFill("solid", fgColor="DDEBF7"),   # blue
    PatternFill("solid", fgColor="E2EFDA"),   # green
    PatternFill("solid", fgColor="FFF2CC"),   # yellow
    PatternFill("solid", fgColor="FCE4EC"),   # pink
    PatternFill("solid", fgColor="FCE5CD"),   # peach
    PatternFill("solid", fgColor="D9F2EF"),   # teal
    PatternFill("solid", fgColor="E8E8F4"),   # grey-lavender
]
_CHARM_SHOP_NEUTRAL_FILL = PatternFill("solid", fgColor="FFFFFF")   # no shop yet


def _charm_shop_fill(shop_name: str, mapping: dict) -> PatternFill:
    """Return a stable background fill for a charm shop.

    Same supplier → same colour. Colours are handed out in first-seen order and
    cycle through ``_CHARM_SHOP_FILLS``; charms with no assigned shop get a
    neutral fill so they are not mistaken for a coloured supplier group.

    ``mapping`` is a per-section dict the caller threads through the row loop so
    each generated section starts its own palette from the top.
    """
    key = (shop_name or "").strip()
    if not key:
        return _CHARM_SHOP_NEUTRAL_FILL
    if key not in mapping:
        mapping[key] = _CHARM_SHOP_FILLS[len(mapping) % len(_CHARM_SHOP_FILLS)]
    return mapping[key]
_CHARM_NA_HDR_FONT  = Font("Calibri", bold=True, color="CCAACC", size=11)  # muted for N/A hdr cells
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP   = Alignment(vertical="center", wrap_text=True)
_CAT_HDR_FILL   = PatternFill("solid", fgColor="2C3E50")
_CAT_HDR_FONT   = Font("Calibri", bold=True, color="FFFFFF", size=12)
_CAT_BODY       = Font("Calibri", size=11)
_CAT_BODY_BOLD  = Font("Calibri", bold=True, size=11)
_CAT_WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")   # amber  – needs attention
_CAT_WARN_FONT  = Font("Calibri", size=10, italic=True, color="7D4E00")
_CAT_PRICE_FILL = PatternFill("solid", fgColor="FFF9E6")   # yellow – price TBD
_CAT_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CAT_WRAP       = Alignment(vertical="center", wrap_text=True)
def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
def _style_row(ws, row: int, cols: int, *, fill=None, font=None) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font      = font or _BODY
        cell.alignment = _WRAP
        cell.border    = _BORDER
        if fill:
            cell.fill = fill
def _embed_photo(ws, photo_bytes: bytes | None, row: int, col: int,
                 photo_px: int = PHOTO_PX) -> None:
    """Embed a JPEG thumbnail anchored to (row, col) of ws."""
    if not photo_bytes:
        return
    try:
        xl_img = XLImage(BytesIO(photo_bytes))
        xl_img.width  = photo_px
        xl_img.height = photo_px
        xl_img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(xl_img)
    except Exception as e:
        log.warning("Photo embed failed at %s%d: %s", get_column_letter(col), row, e)
def _sheet_route(ws, items: list[ResolvedItem],
                 statuses: dict[tuple[str, str], str] | None = None,
                 lang: str = "en",
                 title_fn=None,
                 charm_shops: list[CharmShop] | None = None,
                 charm_library: dict[str, CharmLibraryEntry] | None = None,
                 charm_images_dir: Path | None = None) -> None:
    ws.title = _t("Shopping Route", lang)
    ws.sheet_properties.tabColor = "1F4E79"

    # EN: 15 cols (with Etsy + Private Notes). ZH: compact 8 cols.
    _zh_route_compact = lang == "zh"
    if _zh_route_compact:
        HDRS = [
            "#", _t("Photo", lang), _t("Supplier", lang), _t("Stall", lang),
            _t("Items to Purchase", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Private Notes", lang),
        ]
        COL_ITEMS_TO_PURCHASE = 5
        COL_SUPPLIER, COL_STALL = 3, 4
        COL_PHONE, COL_QTY = 6, 7
        COL_PRIVATE_NOTES = 8
    else:
        HDRS = [
            "#", _t("Photo", lang), _t("Floor", lang), _t("Supplier", lang),
            _t("Stall", lang), _t("Product", lang),
            _t("Items to Purchase", lang),
            _t("Case", lang), _t("Grip", lang), _t("Charm", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Order #", lang),
            _t("Etsy Shop", lang),
            _t("Private Notes", lang),
        ]
        COL_ITEMS_TO_PURCHASE = 7
        COL_CASE, COL_GRIP, COL_CHARM = 8, 9, 10
        COL_SUPPLIER, COL_STALL = 4, 5
        COL_PHONE, COL_QTY = 11, 12
        COL_PRIVATE_NOTES = 15
    COLS    = len(HDRS)
    HDR_ROW = 4
    col_end = get_column_letter(COLS)

    # Use larger row/photo sizing for the Chinese version so images are clearly visible
    _row_h    = ZH_ROW_HEIGHT  if lang == "zh" else ROW_HEIGHT
    _photo_px = ZH_PHOTO_PX    if lang == "zh" else PHOTO_PX

    # -- Title row
    ws.merge_cells(f"A1:{col_end}1")
    if lang == "zh":
        title_date = date.today().strftime("%Y年%m月%d日")
        title_text = f"购物路线  --  {title_date}"
    else:
        title_text = f"Shopping Route  --  {date.today().strftime('%B %d, %Y')}"
    ws.cell(1, 1, title_text).font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    # Three-bucket classification:
    #  • routable    – has supplier with at least shop_name OR stall filled in
    #  • needs_info  – matched a catalog entry whose shop/stall are empty AND
    #                  match score is high-confidence (>= EMPTY_ENTRY_MATCH_THRESHOLD),
    #                  meaning it IS the same product but the user hasn't filled in
    #                  the location yet.  A new amber row was already added to the
    #                  catalog so the user can fill it in.
    #  • unmatched   – no catalog entry, OR a low-confidence match against an
    #                  empty-info entry (potential false positive → also gets its
    #                  own amber row appended to the catalog via update_catalog)
    def _supplier_has_location(r: ResolvedItem) -> bool:
        return bool(r.supplier and (r.supplier.shop_name or r.supplier.stall))

    # Case/grip section lists must exclude items whose style has NO case and NO
    # grip components (i.e. "Charm Only" items).  Those items belong solely in
    # the dedicated Charm section below; including them here would create
    # spurious N/A-only rows that confuse load_items_from_xlsx when recovering
    # orders from an existing Excel file.
    def _needs_casegrip(r: ResolvedItem) -> bool:
        hc, hg, _ = _style_has(r.item.style)
        return hc or hg

    routable   = [r for r in items if _supplier_has_location(r) and _needs_casegrip(r)]
    needs_info = [r for r in items
                  if r.supplier
                  and not _supplier_has_location(r)
                  and not _needs_catalog_entry(r)
                  and _needs_casegrip(r)]
    unmatched  = [r for r in items if (not r.supplier or _needs_catalog_entry(r)) and _needs_casegrip(r)]

    # Charm items: style includes a charm AND the charm is not already bought
    # (Purchased / Out of Production).  Otherwise the Charm section mirrored
    # the whole dashboard instead of the actual procurement list.
    _st = statuses or {}
    charm_items     = [r for r in items if _charm_line_in_shopping_route(r, _st)]
    total_charm_qty = sum(r.item.quantity for r in charm_items)

    supplier_stops = len({(r.supplier.shop_name, r.supplier.stall) for r in routable})
    order_count    = len({r.order.order_number for r in items})

    # -- Subtitle row
    ws.merge_cells(f"A2:{col_end}2")
    if lang == "zh":
        sub_parts = [
            f"{len(items)} 件商品",
            f"{order_count} 个订单",
            f"{supplier_stops} 个供应商",
            "按楼层从低到高排序",
        ]
        if charm_items:
            sub_parts.append(f"{total_charm_qty} 个挂件需采购（独立楼栋）")
        if needs_info:
            sub_parts.append(f"{len(needs_info)} 个待填供应商信息")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} 个未匹配")
    else:
        sub_parts = [
            f"{len(items)} items",
            f"{order_count} orders",
            f"{supplier_stops} supplier stops",
            "sorted lowest to highest floor",
        ]
        if charm_items:
            sub_parts.append(
                f"{total_charm_qty} charm(s) needed \u2014 separate building"
            )
        if needs_info:
            sub_parts.append(f"{len(needs_info)} awaiting supplier info")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} unmatched")
    ws.cell(2, 1, "  |  ".join(sub_parts)).font = _SUB_FONT
    ws.row_dimensions[2].height = 24

    # -- Legend row
    ws.merge_cells(f"A3:{col_end}3")
    if lang == "zh":
        legend_text = (
            "待购项列标明本单需采购的部件（手机壳 / 支架）。"
            "   |   蓝色行 = 已在目录中，请在 supplier_catalog.xlsx 填写供应商信息"
            "   |   ✦ 挂件区（紫色）= 在独立楼栋另行采购，见下方挂件区"
        )
    else:
        legend_text = (
            "Per-component status:   Pending (white)   |   Purchased (green)"
            "   |   Out of Stock (amber)   |   Out of Production (red)"
            "   |   N/A (gray) = not part of this order"
            "   |   blue rows = in catalog, fill supplier info in supplier_catalog.xlsx"
            "   |   \u2728 Charm column = N/A here; charms are purchased at a SEPARATE BUILDING \u2014 see purple section below"
        )
    ws.cell(3, 1, legend_text).font = Font("Calibri", size=9, italic=True, color="555555")
    ws.row_dimensions[3].height = 14

    # -- Header row
    for ci, h in enumerate(HDRS, 1):
        ws.cell(HDR_ROW, ci, h)
    _style_header(ws, HDR_ROW, COLS)
    ws.cell(HDR_ROW, COL_ITEMS_TO_PURCHASE).fill = PatternFill("solid", fgColor="2E7D32")
    if not _zh_route_compact:
        ws.cell(HDR_ROW, COL_CASE).fill  = PatternFill("solid", fgColor="1A6B3C")
        ws.cell(HDR_ROW, COL_GRIP).fill  = PatternFill("solid", fgColor="1A3D6B")
        ws.cell(HDR_ROW, COL_CHARM).fill = PatternFill("solid", fgColor="5B1A6B")
    ws.row_dimensions[HDR_ROW].height = 18

    # -- Group by (supplier, stall), in walking order  (routable only)
    groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in routable:
        groups[(r.supplier.shop_name, r.supplier.stall)].append(r)
    for _gk in groups:
        groups[_gk].sort(key=_route_item_sort_key)

    sorted_keys = sorted(
        groups,
        key=lambda k: _stall_sort_key(k[1], k[0]),
    )

    # Track cells to add to the component-status dropdowns
    active_case_cells:  list[str] = []
    active_grip_cells:  list[str] = []
    active_charm_cells: list[str] = []

    row = HDR_ROW + 1
    first_data_row = row
    seq = 1

    _statuses = statuses or {}

    def _write_component_cell(ws, row, col, has_component, active_cells,
                              order_num="", comp="", item_title=""):
        """Write status (or N/A) into a component cell, restoring preserved values.

        The status key is (order_num, normalized_title, comp) so that items
        from the same order number but with different product titles never
        overwrite each other's statuses.
        """
        cell = ws.cell(row, col)
        if has_component:
            norm_title     = _normalize(item_title)[:50]
            preserved      = _statuses.get((order_num, norm_title, comp))
            cell.value     = _t(preserved, lang) if preserved else _t("Pending", lang)
            cell.alignment = _CENTER
            active_cells.append(cell.coordinate)
        else:
            cell.value     = _t("N/A", lang)
            cell.fill      = _NA_FILL
            cell.font      = _NA_FONT
            cell.alignment = _CENTER

    for gidx, key in enumerate(sorted_keys):
        fill = _GROUP_FILLS[gidx % 2]
        for r in groups[key]:
            floor       = _stall_floor(r.supplier.stall)
            floor_label = f"{floor}F" if floor != 999 else "--"
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            # col 2 = photo
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, r.supplier.shop_name or "--")
                ws.cell(row, COL_STALL, r.supplier.stall or "--")
            else:
                ws.cell(row, 3, floor_label)
                ws.cell(row, 4, r.supplier.shop_name or "--")
                ws.cell(row, 5, r.supplier.stall or "--")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE,  has_case, active_case_cells, onum, "case",  r.item.title)
                _write_component_cell(ws, row, COL_GRIP,  has_grip, active_grip_cells, onum, "grip",  r.item.title)
                # Charm is ALWAYS N/A in case/grip supplier rows — tracked in the
                # dedicated Charm section (separate building) at the bottom of this sheet.
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP

            _style_row(ws, row, COLS, fill=fill)
            if not _zh_route_compact:
                # Re-apply N/A styling after _style_row (which resets font/fill)
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    if not has:
                        c = ws.cell(row, col)
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        ws.cell(row, col).alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP

            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    # -- "Needs Supplier Info" section
    #    Items whose title matched a catalog entry but shop_name and stall are
    #    both empty (amber row in catalog not yet completed by the user).
    #    Shown in pale blue so they are visually distinct from both the main
    #    table and the truly-unmatched amber section below.
    if needs_info:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if lang == "zh":
            ni_banner = (
                "(~)  目录中 \u2013 待填供应商信息  "
                "\u2192  请打开 supplier_catalog.xlsx 并填写店名（D列）和摊位（E列）"
            )
        else:
            ni_banner = (
                "(~)  In Catalog \u2013 Awaiting Supplier Info  "
                "\u2192  open supplier_catalog.xlsx and fill in Shop Name (D) + Stall (E)"
            )
        ni_hdr = ws.cell(row, 1, ni_banner)
        ni_hdr.font   = Font("Calibri", bold=True, size=11, color="1F4E79")
        ni_hdr.fill   = PatternFill("solid", fgColor="BDD7EE")
        ni_hdr.border = _BORDER
        row += 1

        for r in sorted(needs_info, key=_route_item_sort_key):
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, r.supplier.shop_name or "\u2014")
                ws.cell(row, COL_STALL, r.supplier.stall or "\u2014")
            else:
                ws.cell(row, 3, "--")
                ws.cell(row, 4, r.supplier.shop_name or "\u2014")
                ws.cell(row, 5, r.supplier.stall or "\u2014")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE, has_case, active_case_cells, onum, "case", r.item.title)
                _write_component_cell(ws, row, COL_GRIP, has_grip, active_grip_cells, onum, "grip", r.item.title)
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP
            _style_row(ws, row, COLS, fill=_NEEDSINFO_FILL, font=_NEEDSINFO_FONT)
            if not _zh_route_compact:
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    c = ws.cell(row, col)
                    if not has:
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        c.alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    # -- Unmatched items section
    if unmatched:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        warn = ws.cell(
            row, 1,
            "(!!)  未匹配商品 -- 目录中未找到供应商" if lang == "zh"
            else "(!!)  Unmatched Items -- supplier not found in catalog",
        )
        warn.font   = Font("Calibri", bold=True, size=11, color="856404")
        warn.fill   = _WARN_FILL
        warn.border = _BORDER
        row += 1

        for r in sorted(unmatched, key=_route_item_sort_key):
            has_case, has_grip, _ = _style_has(r.item.style)
            onum = r.order.order_number
            norm_title = _normalize(r.item.title)[:50]
            case_status = _statuses.get((onum, norm_title, "case"))
            grip_status = _statuses.get((onum, norm_title, "grip"))
            items_label = _items_to_purchase(has_case, has_grip, case_status, grip_status, lang)

            ws.cell(row, 1, seq)
            if _zh_route_compact:
                ws.cell(row, COL_SUPPLIER, "???")
                ws.cell(row, COL_STALL, "???")
            else:
                ws.cell(row, 3, "--")
                ws.cell(row, 4, "???")
                ws.cell(row, 5, "???")
                ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
            itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, items_label)
            itp_cell.alignment = _CENTER
            itp_cell.font = _ITEMS_TO_PURCHASE_FONT
            if not _zh_route_compact:
                _write_component_cell(ws, row, COL_CASE, has_case, active_case_cells, onum, "case", r.item.title)
                _write_component_cell(ws, row, COL_GRIP, has_grip, active_grip_cells, onum, "grip", r.item.title)
                ws.cell(row, COL_CHARM, _t("N/A", lang))
            ws.cell(row, COL_PHONE, r.item.phone_model)
            ws.cell(row, COL_QTY, r.item.quantity)
            if not _zh_route_compact:
                ws.cell(row, 13, f"#{r.order.order_number}")
                ws.cell(row, 14, r.order.etsy_shop)
            if r.order.private_notes:
                pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                pn.alignment = _WRAP
            _style_row(ws, row, COLS, fill=_WARN_FILL, font=_WARN_FONT)
            if not _zh_route_compact:
                for col, has in (
                    (COL_CASE, has_case), (COL_GRIP, has_grip), (COL_CHARM, False)
                ):
                    c = ws.cell(row, col)
                    if not has:
                        c.fill = _NA_FILL
                        c.font = _NA_FONT
                    else:
                        c.alignment = _CENTER
            if r.order.private_notes:
                ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
            ws.cell(row, 1).alignment = _CENTER
            if not _zh_route_compact:
                ws.cell(row, 3).alignment = _CENTER
            ws.cell(row, COL_QTY).alignment = _CENTER
            ws.row_dimensions[row].height = _row_h
            _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
            seq += 1
            row += 1

    last_data_row = row - 1

    # -- Per-component status dropdowns (only on cells with active components)
    _status_opts = ZH_STATUS_OPTIONS if lang == "zh" else STATUS_OPTIONS
    dv_formula = f'"{",".join(_status_opts)}"'
    dv_kwargs  = dict(
        type="list",
        formula1=dv_formula,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        error="请从下拉列表中选择一个值。" if lang == "zh" else "Pick a value from the dropdown.",
        errorTitle="状态无效" if lang == "zh" else "Invalid status",
    )
    # Case and Grip dropdowns for main supplier sections (charm is in its own section below)
    if not _zh_route_compact:
        for cell_list in (active_case_cells, active_grip_cells):
            if cell_list:
                dv = DataValidation(**dv_kwargs)
                ws.add_data_validation(dv)
                for coord in cell_list:
                    dv.add(coord)

    # -- Conditional formatting: row colour = worst component status
    #
    #    Priority (first rule wins):
    #      1. Any component = Out of Production  -> RED
    #      2. Any component = Out of Stock       -> AMBER  (and none OOP)
    #      3. All included components Purchased  -> GREEN
    #      4. (no rule)                          -> white / group colour (Pending)
    if not _zh_route_compact and last_data_row >= first_data_row:
        full_range = f"A{first_data_row}:{col_end}{last_data_row}"
        r0         = first_data_row  # reference row for relative formulas
        gc = f"${get_column_letter(COL_CASE)}"
        hc = f"${get_column_letter(COL_GRIP)}"
        ic = f"${get_column_letter(COL_CHARM)}"

        oop = _t("Out of Production", lang)
        oos = _t("Out of Stock", lang)
        purchased = _t("Purchased", lang)
        na        = _t("N/A", lang)

        # Rule 1 — any component Out of Production
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[f'OR({gc}{r0}="{oop}",{hc}{r0}="{oop}",{ic}{r0}="{oop}")'],
            fill=_STATUS_FILLS["Out of Production"],
            font=_STATUS_FONTS["Out of Production"],
            stopIfTrue=True,
        ))

        # Rule 2 — any component Out of Stock (and none Out of Production)
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[
                f'AND('
                f'OR({gc}{r0}="{oos}",{hc}{r0}="{oos}",{ic}{r0}="{oos}"),'
                f'NOT(OR({gc}{r0}="{oop}",{hc}{r0}="{oop}",{ic}{r0}="{oop}"))'
                f')'
            ],
            fill=_STATUS_FILLS["Out of Stock"],
            font=_STATUS_FONTS["Out of Stock"],
            stopIfTrue=True,
        ))

        # Rule 3 — all included components Purchased (N/A counts as done)
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[
                f'AND('
                f'OR({gc}{r0}="{na}",{gc}{r0}="{purchased}"),'
                f'OR({hc}{r0}="{na}",{hc}{r0}="{purchased}"),'
                f'OR({ic}{r0}="{na}",{ic}{r0}="{purchased}")'
                f')'
            ],
            fill=_STATUS_FILLS["Purchased"],
            font=_STATUS_FONTS["Purchased"],
            stopIfTrue=True,
        ))

    # ---------------------------------------------------------------------------
    # CHARMS TO PURCHASE — dedicated section, completely separate building
    # ---------------------------------------------------------------------------
    #
    # Two sub-sections:
    #   A) Aggregated by charm code — one row per unique charm, photo + details
    #      from the Charm Library, total qty across all orders.  Status tracked
    #      per charm code via sentinel ``~C:<code>`` in the Order # column.
    #   B) Awaiting charm code — orders whose style has a charm component but
    #      no charm code assigned yet.  Shows product photo + prompt to assign.
    # ---------------------------------------------------------------------------
    if charm_items:
        row += 1   # one blank separator row

        _cshops            = charm_shops or []
        _cshops_lookup_tmp = {cs.shop_name: cs for cs in _cshops}
        total_charm_qty_c  = sum(r.item.quantity for r in charm_items)

        # Partition: items with a charm code vs items still awaiting one
        _coded_items:    list[ResolvedItem] = []
        _awaiting_items: list[ResolvedItem] = []
        for _ci in charm_items:
            _cc = (_ci.supplier.charm_code if _ci.supplier else "").strip()
            if _cc:
                _coded_items.append(_ci)
            else:
                _awaiting_items.append(_ci)

        # Aggregate coded items by charm_code only.  The 1:1 code→shop rule is
        # enforced upstream (normalize_catalog_charm_shops + apply_canonical_
        # charm_fields_to_resolved), so every item with the same code always
        # has the same canonical shop — one row per unique charm code.
        #
        # Shop resolution priority:
        #   1. CharmLibraryEntry.default_charm_shop (canonical ground-truth)
        #   2. r.supplier.charm_shop (fallback for codes not yet in library)
        #   3. "" (empty — row shows "--")
        #
        # Photo resolution is delegated to _charm_photo_for_code so it matches
        # the dashboard exactly: on-disk charm_images/<code>.<ext> FIRST (the
        # live source of truth the dashboard reads + writes), Charm Library BLOB
        # only as a legacy fallback.
        _charm_agg: dict[str, dict] = {}
        for _ci in _coded_items:
            _cc  = _ci.supplier.charm_code.strip()
            _lib = (charm_library or {}).get(_cc)
            # Library default takes precedence — after normalization this is
            # the canonical value.  Fall back to the per-order value only for
            # codes the library hasn't registered yet.
            _cs  = ""
            if _lib and _lib.default_charm_shop:
                _cs = _lib.default_charm_shop.strip()
            if not _cs:
                _cs = (_ci.supplier.charm_shop if _ci.supplier else "").strip()
            if _cc not in _charm_agg:
                # Same source of truth as the dashboard: on-disk image first,
                # Charm Library BLOB fallback (see _charm_photo_for_code).
                _ph = _charm_photo_for_code(_cc, charm_library, charm_images_dir)
                _charm_agg[_cc] = {
                    "code": _cc,
                    "sku": _lib.sku if _lib else "",
                    "default_shop": _lib.default_charm_shop if _lib else "",
                    "notes": _lib.notes if _lib else "",
                    "photo_bytes": _ph,
                    "charm_shop": _cs,
                    "charm_shop_obj": _cshops_lookup_tmp.get(_cs),
                    "total_qty": 0,
                    "orders": [],
                    "items": [],
                    "private_notes": [],  # per-order private notes (deduplicated)
                }
            _charm_agg[_cc]["total_qty"] += _ci.item.quantity
            _charm_agg[_cc]["orders"].append(_ci.order.order_number)
            _charm_agg[_cc]["items"].append(_ci)
            _pn = (_ci.order.private_notes or "").strip()
            if _pn and _pn not in _charm_agg[_cc]["private_notes"]:
                _charm_agg[_cc]["private_notes"].append(_pn)

        n_unique_charms  = len(_charm_agg)
        n_missing_code   = len(_awaiting_items)
        unassigned_count = sum(
            1 for r in charm_items
            if not (r.supplier and r.supplier.charm_shop
                    and r.supplier.charm_shop in _cshops_lookup_tmp)
        )

        # --- Banner row ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if lang == "zh":
            charm_banner_text = (
                f"\u2728  \u5f85\u8d2d\u6302\u4ef6  \u2014  \u72ec\u7acb\u697c\u68cb"
                f"  \u2014  \u5171\u9700 {total_charm_qty_c} \u4e2a\u6302\u4ef6"
                f"\uff0c\u6d89\u53ca {len(charm_items)} \u4e2a\u8ba2\u5355"
            )
            if n_unique_charms:
                charm_banner_text += f"  \u2014  {n_unique_charms} \u79cd\u6302\u4ef6"
            if n_missing_code:
                charm_banner_text += (
                    f"  \u25b6  {n_missing_code} \u4e2a\u8ba2\u5355\u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801"
                    f" \u2014 \u6253\u5f00 supplier_catalog.xlsx \u2192 Product Map H\u5217"
                )
            if unassigned_count:
                charm_banner_text += (
                    f"  \u25b6  {unassigned_count} \u4e2a\u8ba2\u5355\u672a\u5206\u914d\u6302\u4ef6\u5e97"
                )
        else:
            charm_banner_text = (
                f"\u2728  CHARMS TO PURCHASE  \u2014  SEPARATE BUILDING"
                f"  \u2014  {total_charm_qty_c} charm(s) needed"
                f" across {len(charm_items)} order(s)"
            )
            if n_missing_code:
                charm_banner_text += (
                    f"  \u25b6  {n_missing_code} order(s) missing charm-code"
                    f" assignment \u2014 open supplier_catalog.xlsx"
                    f" \u2192 Product Map col H (Charm Code)"
                )
            if unassigned_count:
                charm_banner_text += (
                    f"  \u25b6  {unassigned_count} order(s) missing charm-shop"
                    f" assignment \u2014 open supplier_catalog.xlsx"
                    f" \u2192 Product Map col G (Charm Shop)"
                )
        charm_banner = ws.cell(row, 1, charm_banner_text)
        charm_banner.font      = Font("Calibri", bold=True, size=13, color="FFFFFF")
        charm_banner.fill      = _CHARM_BANNER_FILL
        charm_banner.border    = _BORDER
        charm_banner.alignment = _CENTER
        ws.row_dimensions[row].height = 26
        row += 1

        # --- Charm shops reference row ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if _cshops:
            shop_parts = [f"{s.shop_name}  ({s.stall})" for s in _cshops]
            shops_label = ("\u6302\u4ef6\u5e97\u94fa\uff1a  " if lang == "zh" else "Charm shops:   ")
            shops_text  = shops_label + "   |   ".join(shop_parts)
        else:
            shops_text = (
                "\u672a\u914d\u7f6e\u6302\u4ef6\u5e97\u94fa \u2014 \u8bf7\u5728 supplier_catalog.xlsx \u7684\u6302\u4ef6\u5e97\u94fa\u6807\u7b7e\u4e2d\u6dfb\u52a0"
                if lang == "zh"
                else "No charm shops configured \u2014 add them in the "
                     "'Charm Shops' tab of supplier_catalog.xlsx"
            )
        shops_ref = ws.cell(row, 1, shops_text)
        shops_ref.font      = Font("Calibri", bold=True, size=10, color="3D1359")
        shops_ref.fill      = _CHARM_SHOPS_FILL
        shops_ref.border    = _BORDER
        shops_ref.alignment = _CENTER
        ws.row_dimensions[row].height = 20
        row += 1

        charm_shop_lookup   = _cshops_lookup_tmp
        charm_section_cells: list[str] = []
        charm_first_row     = row

        # ===============================================================
        # SUB-SECTION A — Aggregated charm purchase list (by charm code)
        # ===============================================================
        if _charm_agg:
            if _zh_route_compact:
                _CHARM_HDRS = [
                    "#", _t("Photo", lang),
                    "\u6302\u4ef6\u7f16\u7801",
                    "SKU",
                    "\u6302\u4ef6\u5e97\u94fa",
                    "\u6446\u4f4d",
                    _t("Qty", lang),
                    "\u5907\u6ce8",
                ]
            else:
                _CHARM_HDRS = [
                    "#", _t("Photo", lang), "",
                    "Charm Code" if lang != "zh" else "\u6302\u4ef6\u7f16\u7801",
                    "SKU",
                    "Charm Shop" if lang != "zh" else "\u6302\u4ef6\u5e97\u94fa",
                    _t("Stall", lang),
                    "", "",
                    _t("Charm", lang),
                    "",
                    _t("Qty", lang),
                    "Orders" if lang != "zh" else "\u5173\u8054\u8ba2\u5355",
                    "",
                    "Notes" if lang != "zh" else "\u5907\u6ce8",
                ]
            for ci, h in enumerate(_CHARM_HDRS, 1):
                ws.cell(row, ci, h)
            _style_header(ws, row, COLS)
            if not _zh_route_compact:
                for _muted_col in (3, 8, 9, 11, 14):
                    ws.cell(row, _muted_col).fill = _NA_FILL
                    ws.cell(row, _muted_col).font = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_CHARM).fill = _CHARM_HDR_FILL
            ws.row_dimensions[row].height = 18
            row += 1
            charm_first_row = row

            sorted_codes = sorted(
                _charm_agg,
                key=lambda c: (_charm_agg[c]["charm_shop"] or "\uffff", c),
            )

            _shop_fill_map: dict = {}   # charm shop → stable row tint (this section)
            for cidx, code in enumerate(sorted_codes):
                agg  = _charm_agg[code]

                _cs_name = agg["charm_shop"] or agg["default_shop"]
                # Same supplier → same colour (rows are already grouped by shop).
                fill = _charm_shop_fill(_cs_name, _shop_fill_map)
                _cs_obj  = charm_shop_lookup.get(_cs_name)
                if _cs_obj:
                    shop_display  = _cs_obj.shop_name
                    stall_display = _cs_obj.stall
                elif _cs_name:
                    shop_display  = f"? {_cs_name}"
                    stall_display = "?"
                else:
                    shop_display  = "--"
                    stall_display = "--"

                unique_orders  = sorted(set(agg["orders"]))
                orders_display = ", ".join(f"#{o}" for o in unique_orders)
                # Combine Charm Library notes (static annotation) with the
                # deduplicated per-order private notes collected during aggregation.
                _lib_note    = agg["notes"]
                _order_notes = "; ".join(agg["private_notes"]) if agg["private_notes"] else ""
                if _lib_note and _order_notes:
                    notes_val = f"{_lib_note}\n{_order_notes}"
                else:
                    notes_val = _lib_note or _order_notes

                ws.cell(row, 1, cidx + 1)
                if _zh_route_compact:
                    ws.cell(row, 3, code)
                    ws.cell(row, 4, agg["sku"])
                    ws.cell(row, 5, shop_display)
                    ws.cell(row, 6, stall_display)
                    ws.cell(row, 7, agg["total_qty"])
                    if notes_val:
                        ws.cell(row, 8, notes_val).alignment = _WRAP
                else:
                    ws.cell(row, 3, "")
                    ws.cell(row, 4, code)
                    ws.cell(row, 5, agg["sku"])
                    ws.cell(row, 6, shop_display)
                    ws.cell(row, 7, stall_display)
                    ws.cell(row, 8, "")
                    ws.cell(row, 9, "")

                    preserved = _statuses.get((code, "", "charm_agg"))
                    if not preserved:
                        _per_order = []
                        for _ri in agg["items"]:
                            _ps = _statuses.get((
                                _ri.order.order_number,
                                _normalize(_ri.item.title)[:50],
                                "charm",
                            ))
                            if _ps:
                                _per_order.append(_ps)
                        if _per_order:
                            if any(s == "Out of Production" for s in _per_order):
                                preserved = "Out of Production"
                            elif any(s == "Out of Stock" for s in _per_order):
                                preserved = "Out of Stock"
                            # Only mark the aggregate as Purchased when every single
                            # item has been explicitly marked Purchased.
                            elif (len(_per_order) == len(agg["items"])
                                  and all(s == "Purchased" for s in _per_order)):
                                preserved = "Purchased"

                    charm_cell = ws.cell(row, COL_CHARM)
                    charm_cell.value     = _t(preserved, lang) if preserved else _t("Pending", lang)
                    charm_cell.alignment = _CENTER
                    charm_section_cells.append(charm_cell.coordinate)

                    ws.cell(row, 11, "")
                    ws.cell(row, 12, agg["total_qty"])
                    ws.cell(row, 13, f"~C:{code}")
                    _orders_trunc = orders_display[:60] + ("\u2026" if len(orders_display) > 60 else "")
                    ws.cell(row, 14, _orders_trunc)
                    if notes_val:
                        ws.cell(row, 15, notes_val).alignment = _WRAP

                _style_row(ws, row, COLS, fill=fill)
                if not _zh_route_compact:
                    for _na_c in (3, 8, 9, 11):
                        nc = ws.cell(row, _na_c)
                        nc.fill = _NA_FILL
                        nc.font = _NA_FONT
                    ws.cell(row, COL_CHARM).alignment = _CENTER
                    _sentinel = ws.cell(row, 13)
                    _sentinel.font = Font("Calibri", size=7, color="D8D8D8")
                if notes_val:
                    ws.cell(row, 15 if not _zh_route_compact else 8).alignment = _WRAP
                ws.cell(row, 1).alignment = _CENTER
                ws.cell(row, 4 if not _zh_route_compact else 3).alignment = _CENTER
                ws.cell(row, 12 if not _zh_route_compact else 7).alignment = _CENTER
                ws.row_dimensions[row].height = _row_h
                _embed_photo(ws, agg["photo_bytes"], row, 2, _photo_px)
                row += 1

        charm_last_row = row - 1

        # DataValidation for aggregated charm status cells
        if charm_section_cells:
            dv_charm = DataValidation(**dv_kwargs)
            ws.add_data_validation(dv_charm)
            for coord in charm_section_cells:
                dv_charm.add(coord)

        # Conditional formatting for aggregated charm rows
        if not _zh_route_compact and charm_last_row >= charm_first_row:
            charm_range = f"A{charm_first_row}:{col_end}{charm_last_row}"
            cr0 = charm_first_row
            ic  = f"${get_column_letter(COL_CHARM)}"
            oop = _t("Out of Production", lang)
            oos = _t("Out of Stock", lang)
            purch = _t("Purchased", lang)

            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{oop}"'],
                fill=_STATUS_FILLS["Out of Production"],
                font=_STATUS_FONTS["Out of Production"],
                stopIfTrue=True,
            ))
            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{oos}"'],
                fill=_STATUS_FILLS["Out of Stock"],
                font=_STATUS_FONTS["Out of Stock"],
                stopIfTrue=True,
            ))
            ws.conditional_formatting.add(charm_range, FormulaRule(
                formula=[f'{ic}{cr0}="{purch}"'],
                fill=_STATUS_FILLS["Purchased"],
                font=_STATUS_FONTS["Purchased"],
                stopIfTrue=True,
            ))

        # ===============================================================
        # SUB-SECTION B — Awaiting charm code assignment
        # ===============================================================
        if _awaiting_items:
            row += 1
            _total_await_qty = sum(r.item.quantity for r in _awaiting_items)

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
            if lang == "zh":
                _await_text = (
                    f"\u23f3  \u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801  \u2014  "
                    f"{len(_awaiting_items)} \u4e2a\u8ba2\u5355"
                    f"\uff08{_total_await_qty} \u4ef6\uff09"
                    f"  \u2192  \u6253\u5f00 supplier_catalog.xlsx"
                    f" \u2192 Product Map H\u5217 (\u6302\u4ef6\u7f16\u7801)"
                )
            else:
                _await_text = (
                    f"\u23f3  AWAITING CHARM CODE ASSIGNMENT  \u2014  "
                    f"{len(_awaiting_items)} order(s)"
                    f" ({_total_await_qty} unit(s))"
                    f"  \u2192  open supplier_catalog.xlsx"
                    f" \u2192 Product Map col H (Charm Code)"
                )
            await_cell = ws.cell(row, 1, _await_text)
            await_cell.font      = Font("Calibri", bold=True, size=11, color="7D4E00")
            await_cell.fill      = PatternFill("solid", fgColor="FFF3CD")
            await_cell.border    = _BORDER
            await_cell.alignment = _CENTER
            ws.row_dimensions[row].height = 22
            row += 1

            for ci, h in enumerate(HDRS, 1):
                ws.cell(row, ci, h)
            _style_header(ws, row, COLS)
            if not _zh_route_compact:
                ws.cell(row, COL_CASE).fill  = _NA_FILL
                ws.cell(row, COL_CASE).font  = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_GRIP).fill  = _NA_FILL
                ws.cell(row, COL_GRIP).font  = _CHARM_NA_HDR_FONT
                ws.cell(row, COL_CHARM).fill = PatternFill("solid", fgColor="FFF3CD")
                ws.cell(row, COL_CHARM).font = Font("Calibri", bold=True, size=11, color="7D4E00")
            ws.row_dimensions[row].height = 18
            row += 1

            _AWAIT_FILL      = PatternFill("solid", fgColor="FFFBF0")
            _AWAIT_CODE_FONT = Font("Calibri", size=9, color="7D4E00", italic=True)
            _AWAIT_CODE_FILL = PatternFill("solid", fgColor="FFF3CD")

            def _await_sort(ri: ResolvedItem) -> tuple[str, str]:
                return (_normalize(ri.item.title), ri.order.order_number)

            for aidx, r in enumerate(sorted(_awaiting_items, key=_await_sort)):
                onum = r.order.order_number
                assigned_name = (r.supplier.charm_shop if r.supplier else "") or ""
                assigned_cs   = charm_shop_lookup.get(assigned_name)
                if assigned_cs:
                    shop_display  = assigned_cs.shop_name
                    stall_display = assigned_cs.stall
                elif assigned_name:
                    shop_display  = f"? {assigned_name}"
                    stall_display = "?"
                else:
                    shop_display  = "--"
                    stall_display = "--"

                ws.cell(row, 1, aidx + 1)
                if _zh_route_compact:
                    ws.cell(row, COL_SUPPLIER, shop_display)
                    ws.cell(row, COL_STALL, stall_display)
                else:
                    ws.cell(row, 3, "--")
                    ws.cell(row, 4, shop_display)
                    ws.cell(row, 5, stall_display)
                    ws.cell(row, 6, title_fn(r.item.title) if title_fn else r.item.title)
                itp_cell = ws.cell(row, COL_ITEMS_TO_PURCHASE, "\u2014")
                itp_cell.alignment = _CENTER
                itp_cell.font = _ITEMS_TO_PURCHASE_FONT

                if not _zh_route_compact:
                    for na_col in (COL_CASE, COL_GRIP):
                        nc = ws.cell(row, na_col, _t("N/A", lang))
                        nc.fill      = _NA_FILL
                        nc.font      = _NA_FONT
                        nc.alignment = _CENTER
                    _await_charm = ws.cell(row, COL_CHARM)
                    _await_charm.value     = "\u23f3 Awaiting Code" if lang != "zh" else "\u23f3 \u5f85\u5206\u914d"
                    _await_charm.alignment = _CENTER
                    _await_charm.font      = _AWAIT_CODE_FONT
                    _await_charm.fill      = _AWAIT_CODE_FILL

                ws.cell(row, COL_PHONE, r.item.phone_model)
                ws.cell(row, COL_QTY, r.item.quantity)
                if not _zh_route_compact:
                    ws.cell(row, 13, f"~?#{r.order.order_number}")
                    ws.cell(row, 14, r.order.etsy_shop)
                if r.order.private_notes:
                    pn = ws.cell(row, COL_PRIVATE_NOTES, r.order.private_notes)
                    pn.alignment = _WRAP

                _style_row(ws, row, COLS, fill=_AWAIT_FILL)
                if not _zh_route_compact:
                    for na_col in (COL_CASE, COL_GRIP):
                        nc = ws.cell(row, na_col)
                        nc.fill = _NA_FILL
                        nc.font = _NA_FONT
                    _await_charm = ws.cell(row, COL_CHARM)
                    _await_charm.font = _AWAIT_CODE_FONT
                    _await_charm.fill = _AWAIT_CODE_FILL
                    _await_charm.alignment = _CENTER
                if r.order.private_notes:
                    ws.cell(row, COL_PRIVATE_NOTES).alignment = _WRAP
                ws.cell(row, 1).alignment = _CENTER
                if not _zh_route_compact:
                    ws.cell(row, 3).alignment = _CENTER
                ws.cell(row, COL_QTY).alignment = _CENTER
                ws.row_dimensions[row].height = _row_h
                _embed_photo(ws, r.item.photo_bytes, row, 2, _photo_px)
                row += 1


    # -- Column widths, freeze, filter
    # Product narrowed (26) so Items to Purchase (14) is more prominent
    _photo_col_w = ZH_PHOTO_COL_W if lang == "zh" else PHOTO_COL_W
    if _zh_route_compact:
        col_widths = [4, _photo_col_w, 13, 9, 14, 18, 4, 28]
    else:
        col_widths = [4, _photo_col_w, 6, 13, 9, 26, 14, 10, 10, 10, 18, 4, 16, 15, 32]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{HDR_ROW}:{col_end}{last_data_row}"
def _sheet_orders(ws, items: list[ResolvedItem], lang: str = "en", title_fn=None,
                  charm_library: dict[str, CharmLibraryEntry] | None = None,
                  charm_images_dir: Path | None = None) -> None:
    ws.sheet_properties.tabColor = "2E75B6"

    if lang == "zh":
        HDRS = [
            _t("Buyer", lang), _t("Ship To", lang), _t("Country", lang),
            _t("Order Date", lang), _t("Photo", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Supplier", lang), _t("Stall", lang), _t("Match %", lang),
            _t("Private Notes", lang),
        ]
    else:
        HDRS = [
            _t("Order #", lang), _t("Etsy Shop", lang), _t("Buyer", lang),
            _t("Ship To", lang), _t("Country", lang), _t("Order Date", lang),
            _t("Photo", lang), _t("Product", lang),
            _t("Case", lang), _t("Grip", lang), _t("Charm", lang),
            _t("Phone Model", lang), _t("Qty", lang),
            _t("Supplier", lang), _t("Stall", lang), _t("Match %", lang),
            _t("Private Notes", lang),
        ]
    COLS = len(HDRS)

    for ci, h in enumerate(HDRS, 1):
        ws.cell(1, ci, h)
    _style_header(ws, 1, COLS)
    if lang != "zh":
        _case_col, _grip_col, _charm_col = 9, 10, 11
        ws.cell(1, _case_col).fill  = PatternFill("solid", fgColor="1A6B3C")
        ws.cell(1, _grip_col).fill  = PatternFill("solid", fgColor="1A3D6B")
        ws.cell(1, _charm_col).fill = PatternFill("solid", fgColor="5B1A6B")
    ws.row_dimensions[1].height = 18

    row = 2
    for r in sorted(items, key=lambda x: x.order.order_number):
        if lang == "zh":
            _b, _s, _c, _od, _ph, _pm, _q, _su, _st, _m, _pn = (
                1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11
            )
            ws.cell(row, _b,  r.order.buyer_name)
            ws.cell(row, _s,  r.order.ship_to_name)
            ws.cell(row, _c,  r.order.ship_to_country)
            ws.cell(row, _od, r.order.order_date)
            ws.cell(row, _pm, r.item.phone_model)
            ws.cell(row, _q,  r.item.quantity)
            ws.cell(row, _su, (r.supplier.shop_name or "--") if r.supplier else "--")
            ws.cell(row, _st, (r.supplier.stall or "--")     if r.supplier else "--")
            ws.cell(row, _m,  f"{r.match_score:.0f}%"        if r.supplier else "--")
            if r.order.private_notes:
                ws.cell(row, _pn, r.order.private_notes).alignment = _WRAP
            center_cols = (_q, _m)
        else:
            case, grip, charm = _style_flags(r.item.style)
            ws.cell(row, 1, f"#{r.order.order_number}")
            ws.cell(row, 2, r.order.etsy_shop)
            _b, _s, _c, _od, _ph, _pr, _ca, _g, _ch, _pm, _q, _su, _st, _m, _pn = (
                3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17
            )
            ws.cell(row, _b,  r.order.buyer_name)
            ws.cell(row, _s,  r.order.ship_to_name)
            ws.cell(row, _c,  r.order.ship_to_country)
            ws.cell(row, _od, r.order.order_date)
            ws.cell(row, _pr, title_fn(r.item.title) if title_fn else r.item.title)
            ws.cell(row, _ca, case)
            ws.cell(row, _g,  grip)
            ws.cell(row, _ch, charm)
            ws.cell(row, _pm, r.item.phone_model)
            ws.cell(row, _q,  r.item.quantity)
            ws.cell(row, _su, (r.supplier.shop_name or "--") if r.supplier else "--")
            ws.cell(row, _st, (r.supplier.stall or "--")     if r.supplier else "--")
            ws.cell(row, _m,  f"{r.match_score:.0f}%"        if r.supplier else "--")
            if r.order.private_notes:
                ws.cell(row, _pn, r.order.private_notes).alignment = _WRAP
            center_cols = (_ca, _g, _ch, _q, _m)

        if not r.supplier or _needs_catalog_entry(r):
            fill = _WARN_FILL                          # unmatched or false-positive – amber
        elif not r.supplier.shop_name and not r.supplier.stall:
            fill = _NEEDSINFO_FILL                     # in catalog, info pending – blue
        else:
            fill = None
        _style_row(ws, row, COLS, fill=fill)
        for cc in center_cols:
            ws.cell(row, cc).alignment = _CENTER
        if r.order.private_notes:
            ws.cell(row, _pn).alignment = _WRAP
        ws.row_dimensions[row].height = ROW_HEIGHT
        # The Orders Detail "Photo" column mirrors the Orders Sorting Dashboard's
        # per-order thumbnail, which is ALWAYS the product (phone-case) image —
        # never the charm.  A charm's presence is already conveyed by the ✓ in
        # the Charm column, and the charm's own photo lives in the Shopping Route
        # charm section / Charm Library.  Previously this column showed the charm
        # bead image whenever the style included a charm, producing a photo that
        # contradicted both the dashboard and the row's own Product title (e.g. a
        # "Tamagotchi Clear MagSafe Case" row displaying a bead bracelet).  The
        # product photo is the single source of truth here.
        _embed_photo(ws, r.item.photo_bytes, row, _ph)
        row += 1

    if lang == "zh":
        col_widths = [18, 18, 14, 14, PHOTO_COL_W, 18, 4, 14, 10, 10, 28]
    else:
        col_widths = [16, 18, 16, 18, 14, 14, PHOTO_COL_W, 52, 6, 6, 7, 18, 4, 14, 10, 10, 32]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(COLS)}{row - 1}"
def _sheet_summary(ws, items: list[ResolvedItem], lang: str = "en") -> None:
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:D1")
    ws.cell(1, 1, _t("Summary", lang)).font = _TITLE_FONT
    ws.row_dimensions[1].height = 34

    row = 3
    routable_ct   = sum(1 for r in items
                        if r.supplier and (r.supplier.shop_name or r.supplier.stall))
    needs_info_ct = sum(1 for r in items
                        if r.supplier
                        and not (r.supplier.shop_name or r.supplier.stall)
                        and not _needs_catalog_entry(r))
    unmatched_ct  = sum(1 for r in items if not r.supplier or _needs_catalog_entry(r))
    stats = [
        (_t("Total orders", lang),                              len({r.order.order_number for r in items})),
        (_t("Total line items", lang),                          len(items)),
        (_t("Total quantity", lang),                            sum(r.item.quantity for r in items)),
        (_t("Ready (supplier + location)", lang),               routable_ct),
        (_t("In catalog \u2013 needs supplier info", lang),     needs_info_ct),
        (_t("Not in catalog (unmatched)", lang),                unmatched_ct),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = _BODY_BOLD
        ws.cell(row, 2, val).font   = _BODY
        for c in (1, 2):
            ws.cell(row, c).border = _BORDER
        row += 1

    # Items per supplier (floor-sorted)
    row += 1
    ws.cell(row, 1, _t("Items per Supplier", lang)).font = _SEC_FONT
    row += 1
    for ci, h in enumerate([
        _t("Floor", lang), _t("Supplier", lang), _t("Stall", lang),
        _t("Items", lang), _t("Qty", lang),
    ], 1):
        ws.cell(row, ci, h)
    _style_header(ws, row, 5)
    row += 1

    sup_groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in items:
        if r.supplier:
            sup_groups[(r.supplier.shop_name, r.supplier.stall)].append(r)

    for (shop, stall), grp in sorted(
        sup_groups.items(),
        key=lambda x: _stall_sort_key(x[0][1], x[0][0]),
    ):
        floor = _stall_floor(stall)
        ws.cell(row, 1, f"{floor}F" if floor != 999 else "--")
        ws.cell(row, 2, shop  or "--")
        ws.cell(row, 3, stall or "--")
        ws.cell(row, 4, len(grp))
        ws.cell(row, 5, sum(r.item.quantity for r in grp))
        _style_row(ws, row, 5)
        ws.cell(row, 1).alignment = _CENTER
        row += 1

    # Items per Etsy shop (omit for Chinese — no Etsy shop column)
    if lang != "zh":
        row += 1
        ws.cell(row, 1, _t("Items per Etsy Shop", lang)).font = _SEC_FONT
        row += 1
        for ci, h in enumerate([_t("Etsy Shop", lang), _t("Orders", lang), _t("Items", lang)], 1):
            ws.cell(row, ci, h)
        _style_header(ws, row, 3)
        row += 1

        shop_groups: dict[str, list[ResolvedItem]] = defaultdict(list)
        for r in items:
            shop_groups[r.order.etsy_shop].append(r)

        for shop, grp in sorted(shop_groups.items()):
            ws.cell(row, 1, shop)
            ws.cell(row, 2, len({r.order.order_number for r in grp}))
            ws.cell(row, 3, len(grp))
            _style_row(ws, row, 3)
            row += 1

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 20
def generate_xlsx(items: list[ResolvedItem], output: Path,
                  statuses: dict[tuple[str, str], str] | None = None,
                  lang: str = "en",
                  title_fn=None,
                  charm_shops: list[CharmShop] | None = None,
                  charm_library: dict[str, CharmLibraryEntry] | None = None,
                  charm_images_dir: Path | None = None) -> None:
    wb = openpyxl.Workbook()
    _sheet_route(
        wb.active, items,
        statuses=statuses, lang=lang, title_fn=title_fn,
        charm_shops=charm_shops,
        charm_library=charm_library,
        charm_images_dir=charm_images_dir,
    )
    _sheet_orders(
        wb.create_sheet(_t("Orders Detail", lang)), items,
        lang=lang, title_fn=title_fn, charm_library=charm_library,
        charm_images_dir=charm_images_dir,
    )
    _sheet_summary(wb.create_sheet(_t("Summary", lang)), items, lang=lang)

    # Configure every sheet so File → Export to PDF produces readable output:
    # landscape orientation, fit all columns onto one page wide (unlimited tall),
    # and narrow margins to maximise usable area.
    from openpyxl.worksheet.page import PageMargins
    for ws in wb.worksheets:
        ws.page_setup.orientation        = "landscape"
        ws.page_setup.paperSize          = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage          = True
        ws.page_setup.fitToWidth         = 1
        ws.page_setup.fitToHeight        = 0   # unlimited pages tall
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.6, bottom=0.6,
            header=0.3, footer=0.3,
        )

    wb.save(output)
    log.info("Saved -> %s", output.resolve())
def _sheet_route_simple(
    ws,
    items: list[ResolvedItem],
    statuses: dict[tuple[str, str], str] | None = None,
    charm_shops: list[CharmShop] | None = None,
    charm_library: dict[str, CharmLibraryEntry] | None = None,
    charm_images_dir: Path | None = None,
    lang: str = "en",
    show_components: bool | None = None,
) -> None:
    ws.title = _t("Shopping Route", lang)
    ws.sheet_properties.tabColor = "1F4E79"

    # Chinese version omits the Private Notes column (col 10) to keep the
    # sheet compact — private notes are English-only buyer messages.
    #
    # Case / Grip status columns (cols 6-7):
    #   • English  → always shown.
    #   • Chinese  → hidden by default (the compact 待购项 column is enough),
    #                but forced ON for the extra "Chinese + status" workbook
    #                via show_components=True, so the purchasing employee can
    #                record the per-component buy status (Pending / Purchased /
    #                Out of Stock / Out of Production) directly — mirroring the
    #                full route. The Charm status column lives in Section 2 and
    #                already carries the same dropdown in both languages.
    _has_notes_col  = lang != "zh"
    _has_comp_cols  = (lang != "zh") if show_components is None else show_components
    if _has_notes_col:
        COLS = 10
    elif _has_comp_cols:
        COLS = 9
    else:
        COLS = 7
    col_end = get_column_letter(COLS)
    HDR_ROW = 4

    # ── Section 1 column positions ───────────────────────────────────────
    S1_PHOTO    = 2
    S1_SUPPLIER = 3
    S1_STALL    = 4
    S1_ITP      = 5   # Items to Purchase
    S1_CASE     = 6               if _has_comp_cols else None
    S1_GRIP     = 7               if _has_comp_cols else None
    S1_PHONE    = 8               if _has_comp_cols else 6
    S1_QTY      = 9               if _has_comp_cols else 7
    S1_NOTES    = 10 if _has_notes_col else None   # None → column omitted

    # ── Section 2 column positions (same grid, different semantics) ───────
    S2_PHOTO       = 2
    S2_CHARM_CODE  = 3
    S2_CHARM_SHOP  = 4
    S2_STALL       = 5
    S2_CHARM       = 6
    # cols 7-8 intentionally blank / N/A (only when comp cols present)
    S2_QTY         = 9 if _has_comp_cols else 7
    S2_NOTES       = 10 if _has_notes_col else None   # None → column omitted

    # Hidden round-trip key column (only on the per-component status workbook —
    # the file the purchasing employee edits and re-uploads). Lives one column
    # past the last visible column so it never disturbs the layout, merges,
    # auto-filter or print area; it is hidden + zero-width at the bottom.
    _emit_key = bool(_has_comp_cols)
    KEY_COL   = (COLS + 1) if _emit_key else None

    _statuses  = statuses or {}
    _row_h     = ROW_HEIGHT
    _photo_px  = PHOTO_PX
    _photo_col_w = PHOTO_COL_W

    # ── Classify items ────────────────────────────────────────────────────
    def _has_loc(r: ResolvedItem) -> bool:
        return bool(r.supplier and (r.supplier.shop_name or r.supplier.stall))

    routable   = [r for r in items if _has_loc(r)]
    needs_info = [r for r in items if r.supplier and not _has_loc(r) and not _needs_catalog_entry(r)]
    unmatched  = [r for r in items if not r.supplier or _needs_catalog_entry(r)]
    charm_items = [r for r in items if _charm_line_in_shopping_route(r, _statuses)]

    total_charm_qty = sum(r.item.quantity for r in charm_items)
    supplier_stops  = len({(r.supplier.shop_name, r.supplier.stall) for r in routable})
    order_count     = len({r.order.order_number for r in items})

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_end}1")
    _title_date = (
        date.today().strftime("%Y\u5e74%m\u6708%d\u65e5")
        if lang == "zh" else
        date.today().strftime("%B %d, %Y")
    )
    ws.cell(1, 1, f"{_t('Shopping Route', lang)}  --  {_title_date}").font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    # ── Subtitle ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{col_end}2")
    if lang == "zh":
        sub_parts = [
            f"{len(items)} \u4ef6\u5546\u54c1",
            f"{order_count} \u4e2a\u8ba2\u5355",
            f"{supplier_stops} \u4e2a\u4f9b\u5e94\u5546",
            _t("sorted lowest to highest floor", lang),
        ]
        if charm_items:
            sub_parts.append(f"{total_charm_qty} \u4e2a\u6302\u4ef6\u9700\u91c7\u8d2d \u2014 \u72ec\u7acb\u697c\u68cb")
        if needs_info:
            sub_parts.append(f"{len(needs_info)} \u4e2a{_t('awaiting supplier info', lang)}")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} \u4e2a{_t('unmatched', lang)}")
    else:
        sub_parts = [
            f"{len(items)} items",
            f"{order_count} orders",
            f"{supplier_stops} supplier stops",
            _t("sorted lowest to highest floor", lang),
        ]
        if charm_items:
            sub_parts.append(f"{total_charm_qty} charm(s) needed \u2014 separate building")
        if needs_info:
            sub_parts.append(f"{len(needs_info)} awaiting supplier info")
        if unmatched:
            sub_parts.append(f"{len(unmatched)} unmatched")
    ws.cell(2, 1, "  |  ".join(sub_parts)).font = _SUB_FONT
    ws.row_dimensions[2].height = 24

    # ── Legend ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A3:{col_end}3")
    _legend_key = "Status legend: Pending (white) | Purchased (green) | Out of Stock (amber) | Out of Production (red) | N/A (gray)"
    ws.cell(3, 1, _t(_legend_key, lang) if lang == "zh" else (
        "Status:  Pending (white)   |   Purchased (green)"
        "   |   Out of Stock (amber)   |   Out of Production (red)"
        "   |   N/A (gray)"
    )).font = Font("Calibri", size=9, italic=True, color="555555")
    ws.row_dimensions[3].height = 14

    # ── Section 1 header row ──────────────────────────────────────────────
    S1_HDRS = [
        "#",
        _t("Photo", lang),
        _t("Supplier", lang),
        _t("Stall", lang),
        _t("Items to Purchase", lang),
        *([_t("Case", lang), _t("Grip", lang)] if _has_comp_cols else []),
        _t("Phone Model", lang),
        _t("Qty", lang),
        *([_t("Private Notes", lang)] if _has_notes_col else []),
    ]
    for ci, h in enumerate(S1_HDRS, 1):
        ws.cell(HDR_ROW, ci, h)
    _style_header(ws, HDR_ROW, COLS)
    if KEY_COL is not None:
        ws.cell(HDR_ROW, KEY_COL, _UED_KEY_HEADER)
    ws.cell(HDR_ROW, S1_ITP).fill = PatternFill("solid", fgColor="2E7D32")
    if _has_comp_cols:
        ws.cell(HDR_ROW, S1_CASE).fill = PatternFill("solid", fgColor="1A6B3C")
        ws.cell(HDR_ROW, S1_GRIP).fill = PatternFill("solid", fgColor="1A3D6B")
    ws.row_dimensions[HDR_ROW].height = 18

    # ── Group routable items by supplier ──────────────────────────────────
    groups: dict[tuple[str, str], list[ResolvedItem]] = defaultdict(list)
    for r in routable:
        groups[(r.supplier.shop_name, r.supplier.stall)].append(r)
    for gk in groups:
        groups[gk].sort(key=_route_item_sort_key)
    sorted_keys = sorted(
        groups,
        key=lambda k: _stall_sort_key(k[1], k[0]),
    )

    active_case_cells: list[str] = []
    active_grip_cells: list[str] = []
    row = HDR_ROW + 1
    first_data_row = row
    seq = 1

    _status_opts = ZH_STATUS_OPTIONS if lang == "zh" else STATUS_OPTIONS
    dv_formula   = f'"{",".join(_status_opts)}"'
    dv_kwargs    = dict(
        type="list", formula1=dv_formula, allow_blank=False, showDropDown=False,
        showErrorMessage=True,
        error=("请从下拉列表中选择一个值。" if lang == "zh"
               else "Pick a value from the dropdown."),
        errorTitle=("状态无效" if lang == "zh" else "Invalid status"),
    )

    def _write_comp(ws, row, col, has_component, active_cells, order_num="", comp="", item_title=""):
        cell = ws.cell(row, col)
        if has_component:
            nt  = _normalize(item_title)[:50]
            prv = _statuses.get((order_num, nt, comp))
            cell.value     = _t(prv, lang) if prv else _t("Pending", lang)
            cell.alignment = _CENTER
            active_cells.append(cell.coordinate)
        else:
            cell.value     = _t("N/A", lang)
            cell.fill      = _NA_FILL
            cell.font      = _NA_FONT
            cell.alignment = _CENTER

    def _write_s1_row(ws, row, r: ResolvedItem, fill, supplier_display, stall_display):
        nonlocal seq
        has_case, has_grip, _ = _style_has(r.item.style)
        onum = r.order.order_number
        nt   = _normalize(r.item.title)[:50]
        case_st = _statuses.get((onum, nt, "case"))
        grip_st = _statuses.get((onum, nt, "grip"))
        items_label = _items_to_purchase(has_case, has_grip, case_st, grip_st, lang)

        ws.cell(row, 1, seq)
        ws.cell(row, S1_SUPPLIER, supplier_display)
        ws.cell(row, S1_STALL, stall_display)
        itp = ws.cell(row, S1_ITP, items_label)
        itp.alignment = _CENTER
        itp.font      = _ITEMS_TO_PURCHASE_FONT
        if S1_CASE is not None:
            _write_comp(ws, row, S1_CASE, has_case, active_case_cells, onum, "case", r.item.title)
        if S1_GRIP is not None:
            _write_comp(ws, row, S1_GRIP, has_grip, active_grip_cells, onum, "grip", r.item.title)
        ws.cell(row, S1_PHONE, r.item.phone_model)
        ws.cell(row, S1_QTY,   r.item.quantity)
        if S1_NOTES and r.order.private_notes:
            ws.cell(row, S1_NOTES, r.order.private_notes).alignment = _WRAP

        _style_row(ws, row, COLS, fill=fill)
        if S1_CASE is not None:
            for ci, has in ((S1_CASE, has_case), (S1_GRIP, has_grip)):
                c = ws.cell(row, ci)
                if not has:
                    c.fill = _NA_FILL
                    c.font = _NA_FONT
                else:
                    c.alignment = _CENTER
        if S1_NOTES and r.order.private_notes:
            ws.cell(row, S1_NOTES).alignment = _WRAP
        ws.cell(row, 1).alignment      = _CENTER
        ws.cell(row, S1_QTY).alignment = _CENTER
        ws.row_dimensions[row].height  = _row_h
        if KEY_COL is not None:
            ws.cell(row, KEY_COL, json.dumps(
                {"t": "l",
                 "r": str(r.order.order_number).strip(),
                 "k": _ued_line_item_key(r.item.title, r.item.listing_id)},
                ensure_ascii=False, separators=(",", ":"),
            ))
        _embed_photo(ws, r.item.photo_bytes, row, S1_PHOTO, _photo_px)
        seq += 1

    # ── Chinese simple: pre-filter completed / charm-only rows out of every ──
    # section list BEFORE any writing starts.  This is done once at the data  ─
    # level so all three loops below need no extra guards.                     ─
    #                                                                          ─
    # Excluded when lang == "zh":                                              ─
    #   • Charm-only items (no case, no grip) — they live in Section 2.       ─
    #   • Items whose 待购项 would be "—": all case/grip components are        ─
    #     already Purchased (Out-of-Stock / Out-of-Production items remain     ─
    #     visible because employees still need to act on those).               ─
    if lang == "zh":
        def _s1_needed(r: ResolvedItem) -> bool:
            hc, hg, _ = _style_has(r.item.style)
            if not hc and not hg:
                return False        # charm-only → belongs in Section 2 only
            _onum = r.order.order_number
            _nt   = _normalize(r.item.title)[:50]
            return _items_to_purchase(
                hc, hg,
                _statuses.get((_onum, _nt, "case")),
                _statuses.get((_onum, _nt, "grip")),
                "en",               # compare against the fixed "—" em-dash
            ) != "\u2014"

        for _k in list(sorted_keys):
            groups[_k] = [r for r in groups[_k] if _s1_needed(r)]
        sorted_keys = [k for k in sorted_keys if groups[k]]
        needs_info  = [r for r in needs_info  if _s1_needed(r)]
        unmatched   = [r for r in unmatched   if _s1_needed(r)]

    # ── Routable rows ─────────────────────────────────────────────────────
    for gidx, key in enumerate(sorted_keys):
        fill = _GROUP_FILLS[gidx % 2]
        for r in groups[key]:
            _write_s1_row(ws, row, r, fill,
                          r.supplier.shop_name or "--",
                          r.supplier.stall     or "--")
            row += 1

    # ── Needs Supplier Info ───────────────────────────────────────────────
    if needs_info:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        _ni_label = _t("In Catalog \u2013 Awaiting Supplier Info", lang)
        _ni_action = _t("open supplier_catalog.xlsx and fill in Shop Name + Stall", lang)
        ni = ws.cell(row, 1, f"(~)  {_ni_label}  \u2192  {_ni_action}")
        ni.font   = Font("Calibri", bold=True, size=11, color="1F4E79")
        ni.fill   = PatternFill("solid", fgColor="BDD7EE")
        ni.border = _BORDER
        row += 1
        for r in sorted(needs_info, key=_route_item_sort_key):
            _write_s1_row(ws, row, r, _NEEDSINFO_FILL,
                          r.supplier.shop_name or "\u2014",
                          r.supplier.stall     or "\u2014")
            # Re-apply needs-info font to non-N/A cells after _write_s1_row
            for ci in range(1, COLS + 1):
                c = ws.cell(row, ci)
                if c.fill.fgColor.value != "EFEFEF":   # skip N/A cells
                    c.font = _NEEDSINFO_FONT
            row += 1

    # ── Unmatched ─────────────────────────────────────────────────────────
    if unmatched:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        _um_label = _t("Unmatched Items \u2014 supplier not found in catalog", lang)
        warn = ws.cell(row, 1, f"(!!)  {_um_label}")
        warn.font   = Font("Calibri", bold=True, size=11, color="856404")
        warn.fill   = _WARN_FILL
        warn.border = _BORDER
        row += 1
        for r in sorted(unmatched, key=_route_item_sort_key):
            _write_s1_row(ws, row, r, _WARN_FILL, "???", "???")
            for ci in range(1, COLS + 1):
                ws.cell(row, ci).font = _WARN_FONT
            row += 1

    last_data_row = row - 1

    # ── Status dropdowns for Case / Grip ──────────────────────────────────
    for cell_list in (active_case_cells, active_grip_cells):
        if cell_list:
            dv = DataValidation(**dv_kwargs)
            ws.add_data_validation(dv)
            for coord in cell_list:
                dv.add(coord)

    # ── Conditional formatting for Section 1 ──────────────────────────────
    # Use the localised status strings that were written into cells above.
    # Skipped for Chinese (no Case/Grip columns to drive row colouring).
    _sv_oop  = _t("Out of Production", lang)
    _sv_oos  = _t("Out of Stock", lang)
    _sv_done = _t("Purchased", lang)
    _sv_na   = _t("N/A", lang)
    if _has_comp_cols and last_data_row >= first_data_row:
        s1_range = f"A{first_data_row}:{col_end}{last_data_row}"
        r0 = first_data_row
        gc = f"${get_column_letter(S1_CASE)}"
        hc = f"${get_column_letter(S1_GRIP)}"
        ws.conditional_formatting.add(s1_range, FormulaRule(
            formula=[f'OR({gc}{r0}="{_sv_oop}",{hc}{r0}="{_sv_oop}")'],
            fill=_STATUS_FILLS["Out of Production"], font=_STATUS_FONTS["Out of Production"], stopIfTrue=True,
        ))
        ws.conditional_formatting.add(s1_range, FormulaRule(
            formula=[f'AND(OR({gc}{r0}="{_sv_oos}",{hc}{r0}="{_sv_oos}"),'
                     f'NOT(OR({gc}{r0}="{_sv_oop}",{hc}{r0}="{_sv_oop}")))'],
            fill=_STATUS_FILLS["Out of Stock"], font=_STATUS_FONTS["Out of Stock"], stopIfTrue=True,
        ))
        ws.conditional_formatting.add(s1_range, FormulaRule(
            formula=[f'AND(OR({gc}{r0}="{_sv_na}",{gc}{r0}="{_sv_done}"),'
                     f'OR({hc}{r0}="{_sv_na}",{hc}{r0}="{_sv_done}"))'],
            fill=_STATUS_FILLS["Purchased"], font=_STATUS_FONTS["Purchased"], stopIfTrue=True,
        ))

    # =========================================================================
    # SECTION 2 — CHARMS (aggregated by charm code)
    # =========================================================================
    if charm_items:
        row += 1

        _cshops        = charm_shops or []
        _cshops_lookup = {cs.shop_name: cs for cs in _cshops}
        total_charm_qty_c = sum(r.item.quantity for r in charm_items)

        # Partition coded vs awaiting
        _coded_items:    list[ResolvedItem] = []
        _awaiting_items: list[ResolvedItem] = []
        for _ci in charm_items:
            _cc = (_ci.supplier.charm_code if _ci.supplier else "").strip()
            (_coded_items if _cc else _awaiting_items).append(_ci)

        # Aggregate by charm_code — one row per unique charm.  The 1:1 rule is
        # enforced upstream so every item with the same code already has the
        # same canonical shop.  Photo and shop resolution match _sheet_route
        # exactly: Charm Library first, folder/per-order fallback.
        _charm_agg: dict[str, dict] = {}
        for _ci in _coded_items:
            _cc  = _ci.supplier.charm_code.strip()
            _lib = (charm_library or {}).get(_cc)
            _cs  = ""
            if _lib and _lib.default_charm_shop:
                _cs = _lib.default_charm_shop.strip()
            if not _cs:
                _cs = (_ci.supplier.charm_shop if _ci.supplier else "").strip()
            if _cc not in _charm_agg:
                # Same source of truth as the dashboard: on-disk image first,
                # Charm Library BLOB fallback (see _charm_photo_for_code).
                _ph = _charm_photo_for_code(_cc, charm_library, charm_images_dir)
                _charm_agg[_cc] = {
                    "code": _cc,
                    "default_shop": _lib.default_charm_shop if _lib else "",
                    "photo_bytes": _ph,
                    "charm_shop": _cs,
                    "total_qty": 0,
                    "orders": [],
                    "items": [],
                }
            _charm_agg[_cc]["total_qty"] += _ci.item.quantity
            _charm_agg[_cc]["orders"].append(_ci.order.order_number)
            _charm_agg[_cc]["items"].append(_ci)

        n_missing_code   = len(_awaiting_items)
        unassigned_count = sum(
            1 for r in charm_items
            if not (r.supplier and r.supplier.charm_shop and r.supplier.charm_shop in _cshops_lookup)
        )

        # ── Charm banner ──────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if lang == "zh":
            cb_text = (
                f"\u2728  \u5f85\u8d2d\u6302\u4ef6  \u2014  \u72ec\u7acb\u697c\u68cb"
                f"  \u2014  {total_charm_qty_c} \u4e2a\u6302\u4ef6\uff0c\u6d89\u53ca {len(charm_items)} \u4e2a\u8ba2\u5355"
            )
            if n_missing_code:
                cb_text += f"  \u25b6  {n_missing_code} \u4e2a\u8ba2\u5355\u7f3a\u6302\u4ef6\u7f16\u7801"
            if unassigned_count:
                cb_text += f"  \u25b6  {unassigned_count} \u4e2a\u8ba2\u5355\u7f3a\u6302\u4ef6\u5e97\u94fa"
        else:
            cb_text = (
                f"\u2728  CHARMS TO PURCHASE  \u2014  SEPARATE BUILDING"
                f"  \u2014  {total_charm_qty_c} charm(s) needed across {len(charm_items)} order(s)"
            )
            if n_missing_code:
                cb_text += f"  \u25b6  {n_missing_code} order(s) missing charm-code assignment"
            if unassigned_count:
                cb_text += f"  \u25b6  {unassigned_count} order(s) missing charm-shop assignment"
        cb = ws.cell(row, 1, cb_text)
        cb.font      = Font("Calibri", bold=True, size=13, color="FFFFFF")
        cb.fill      = _CHARM_BANNER_FILL
        cb.border    = _BORDER
        cb.alignment = _CENTER
        ws.row_dimensions[row].height = 26
        row += 1

        # ── Charm shops reference ─────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        if _cshops:
            _shops_label = "\u6302\u4ef6\u5e97\u94fa\uff1a   " if lang == "zh" else "Charm shops:   "
            shops_text = _shops_label + "   |   ".join(
                f"{s.shop_name}  ({s.stall})" for s in _cshops
            )
        else:
            shops_text = _t("No charm shops configured \u2014 add them in the \u2018Charm Shops\u2019 tab", lang)
        sr = ws.cell(row, 1, shops_text)
        sr.font      = Font("Calibri", bold=True, size=10, color="3D1359")
        sr.fill      = _CHARM_SHOPS_FILL
        sr.border    = _BORDER
        sr.alignment = _CENTER
        ws.row_dimensions[row].height = 20
        row += 1

        # ── Section 2 sub-header ─────────────────────────────────────────
        S2_HDRS = [
            "#",
            _t("Photo", lang),
            _t("Charm Code", lang),
            _t("Charm Shop", lang),
            _t("Stall", lang),
            _t("Charm", lang),
            *(["", ""] if _has_comp_cols else []),
            _t("Qty", lang),
            *([_t("Private Notes", lang)] if _has_notes_col else []),
        ]
        for ci, h in enumerate(S2_HDRS, 1):
            ws.cell(row, ci, h)
        _style_header(ws, row, COLS)
        if KEY_COL is not None:
            ws.cell(row, KEY_COL, _UED_KEY_HEADER)
        ws.cell(row, S2_CHARM).fill = _CHARM_HDR_FILL
        if _has_comp_cols:
            for _mc in (7, 8):
                ws.cell(row, _mc).fill = _NA_FILL
                ws.cell(row, _mc).font = _CHARM_NA_HDR_FONT
        ws.row_dimensions[row].height = 18
        row += 1

        charm_first_row     = row
        charm_section_cells: list[str] = []

        # ── Aggregated charm rows ─────────────────────────────────────────
        sorted_codes = sorted(
            _charm_agg,
            key=lambda c: (_charm_agg[c]["charm_shop"] or "\uffff", c),
        )
        _shop_fill_map: dict = {}   # charm shop → stable row tint (this section)
        for cidx, code in enumerate(sorted_codes):
            agg  = _charm_agg[code]

            _cs_name = agg["charm_shop"] or agg["default_shop"]
            # Same supplier → same colour (rows are already grouped by shop).
            fill = _charm_shop_fill(_cs_name, _shop_fill_map)
            _cs_obj  = _cshops_lookup.get(_cs_name)
            if _cs_obj:
                shop_disp  = _cs_obj.shop_name
                stall_disp = _cs_obj.stall
            elif _cs_name:
                shop_disp  = f"? {_cs_name}"
                stall_disp = "?"
            else:
                shop_disp  = "--"
                stall_disp = "--"

            preserved = _statuses.get((code, "", "charm_agg"))
            if not preserved:
                _per_order = []
                for _ri in agg["items"]:
                    _ps = _statuses.get((_ri.order.order_number, _normalize(_ri.item.title)[:50], "charm"))
                    if _ps:
                        _per_order.append(_ps)
                if _per_order:
                    if any(s == "Out of Production" for s in _per_order):
                        preserved = "Out of Production"
                    elif any(s == "Out of Stock" for s in _per_order):
                        preserved = "Out of Stock"
                    # Only mark the aggregate as Purchased when every single item
                    # in the group has been explicitly marked Purchased.  A partial
                    # subset of "Purchased" statuses must not hide a charm that still
                    # has outstanding (implicitly-Pending) orders.
                    elif (len(_per_order) == len(agg["items"])
                          and all(s == "Purchased" for s in _per_order)):
                        preserved = "Purchased"

            # Chinese simple: hide charms that are already Purchased.
            if lang == "zh" and preserved == "Purchased":
                continue

            # Collect private notes from all orders (deduplicated)
            _all_notes: list[str] = []
            _seen_notes: set[str] = set()
            for _ri in agg["items"]:
                _pn = (_ri.order.private_notes or "").strip()
                if _pn and _pn not in _seen_notes:
                    _all_notes.append(_pn)
                    _seen_notes.add(_pn)

            ws.cell(row, 1, cidx + 1)
            ws.cell(row, S2_CHARM_CODE, code)
            ws.cell(row, S2_CHARM_SHOP, shop_disp)
            ws.cell(row, S2_STALL,      stall_disp)
            charm_cell = ws.cell(row, S2_CHARM)
            charm_cell.value     = _t(preserved, lang) if preserved else _t("Pending", lang)
            charm_cell.alignment = _CENTER
            charm_section_cells.append(charm_cell.coordinate)
            ws.cell(row, S2_QTY, agg["total_qty"])
            if S2_NOTES and _all_notes:
                ws.cell(row, S2_NOTES, "; ".join(_all_notes)).alignment = _WRAP

            _style_row(ws, row, COLS, fill=fill)
            if _has_comp_cols:
                for _mc in (7, 8):
                    ws.cell(row, _mc).fill = _NA_FILL
                    ws.cell(row, _mc).font = _NA_FONT
            ws.cell(row, S2_CHARM).alignment      = _CENTER
            ws.cell(row, 1).alignment             = _CENTER
            ws.cell(row, S2_CHARM_CODE).alignment = _CENTER
            ws.cell(row, S2_QTY).alignment        = _CENTER
            if S2_NOTES and _all_notes:
                ws.cell(row, S2_NOTES).alignment = _WRAP
            ws.row_dimensions[row].height = _row_h
            if KEY_COL is not None:
                # Carry every constituent line (order + line key + qty) so the
                # importer maps this single aggregated charm status back onto the
                # exact route_assignments rows it represents — including
                # charm-only orders that never appear in Section 1.
                _constituents = [
                    [str(_ri.order.order_number).strip(),
                     _ued_line_item_key(_ri.item.title, _ri.item.listing_id),
                     int(_ri.item.quantity or 1)]
                    for _ri in agg["items"]
                ]
                ws.cell(row, KEY_COL, json.dumps(
                    {"t": "c", "code": code, "lines": _constituents},
                    ensure_ascii=False, separators=(",", ":"),
                ))
            _embed_photo(ws, agg["photo_bytes"], row, S2_PHOTO, _photo_px)
            row += 1

        charm_last_row = row - 1

        # ── Charm status dropdowns ────────────────────────────────────────
        if charm_section_cells:
            dv_charm = DataValidation(**dv_kwargs)
            ws.add_data_validation(dv_charm)
            for coord in charm_section_cells:
                dv_charm.add(coord)

        # ── Conditional formatting for Section 2 ──────────────────────────
        # Use the same localised status strings as were written into cells.
        if charm_last_row >= charm_first_row:
            c2_range = f"A{charm_first_row}:{col_end}{charm_last_row}"
            ic       = f"${get_column_letter(S2_CHARM)}"
            cr0      = charm_first_row
            ws.conditional_formatting.add(c2_range, FormulaRule(
                formula=[f'{ic}{cr0}="{_sv_oop}"'],
                fill=_STATUS_FILLS["Out of Production"], font=_STATUS_FONTS["Out of Production"], stopIfTrue=True,
            ))
            ws.conditional_formatting.add(c2_range, FormulaRule(
                formula=[f'{ic}{cr0}="{_sv_oos}"'],
                fill=_STATUS_FILLS["Out of Stock"], font=_STATUS_FONTS["Out of Stock"], stopIfTrue=True,
            ))
            ws.conditional_formatting.add(c2_range, FormulaRule(
                formula=[f'{ic}{cr0}="{_sv_done}"'],
                fill=_STATUS_FILLS["Purchased"], font=_STATUS_FONTS["Purchased"], stopIfTrue=True,
            ))

        # ── Awaiting charm code sub-section ──────────────────────────────
        if _awaiting_items:
            row += 1
            _total_await_qty = sum(r.item.quantity for r in _awaiting_items)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
            if lang == "zh":
                _aw_action = _t("open supplier_catalog.xlsx \u2192 Product Map col H (Charm Code)", lang)
                _aw_text = (
                    f"\u23f3  \u5f85\u5206\u914d\u6302\u4ef6\u7f16\u7801  \u2014  "
                    f"{len(_awaiting_items)} \u4e2a\u8ba2\u5355\uff08{_total_await_qty} \u4ef6\uff09"
                    f"  \u2192  {_aw_action}"
                )
            else:
                _aw_text = (
                    f"\u23f3  AWAITING CHARM CODE ASSIGNMENT  \u2014  "
                    f"{len(_awaiting_items)} order(s) ({_total_await_qty} unit(s))"
                    f"  \u2192  open supplier_catalog.xlsx \u2192 Product Map col H (Charm Code)"
                )
            aw = ws.cell(row, 1, _aw_text)
            aw.font      = Font("Calibri", bold=True, size=11, color="7D4E00")
            aw.fill      = PatternFill("solid", fgColor="FFF3CD")
            aw.border    = _BORDER
            aw.alignment = _CENTER
            ws.row_dimensions[row].height = 22
            row += 1

            for ci, h in enumerate(S2_HDRS, 1):
                ws.cell(row, ci, h)
            _style_header(ws, row, COLS)
            ws.cell(row, S2_CHARM).fill = PatternFill("solid", fgColor="FFF3CD")
            ws.cell(row, S2_CHARM).font = Font("Calibri", bold=True, size=11, color="7D4E00")
            if _has_comp_cols:
                for _mc in (7, 8):
                    ws.cell(row, _mc).fill = _NA_FILL
                    ws.cell(row, _mc).font = _CHARM_NA_HDR_FONT
            ws.row_dimensions[row].height = 18
            row += 1

            _AWAIT_FILL      = PatternFill("solid", fgColor="FFFBF0")
            _AWAIT_CODE_FONT = Font("Calibri", size=9, color="7D4E00", italic=True)
            _AWAIT_CODE_FILL = PatternFill("solid", fgColor="FFF3CD")

            for aidx, r in enumerate(sorted(
                _awaiting_items,
                key=lambda ri: (_normalize(ri.item.title), ri.order.order_number),
            )):
                assigned_name = (r.supplier.charm_shop if r.supplier else "") or ""
                assigned_cs   = _cshops_lookup.get(assigned_name)
                if assigned_cs:
                    shop_disp  = assigned_cs.shop_name
                    stall_disp = assigned_cs.stall
                elif assigned_name:
                    shop_disp  = f"? {assigned_name}"
                    stall_disp = "?"
                else:
                    shop_disp  = "--"
                    stall_disp = "--"

                ws.cell(row, 1, aidx + 1)
                ws.cell(row, S2_CHARM_CODE, "\u2014")
                ws.cell(row, S2_CHARM_SHOP, shop_disp)
                ws.cell(row, S2_STALL,      stall_disp)
                _ac = ws.cell(row, S2_CHARM, _t("\u23f3 Awaiting Code", lang))
                _ac.alignment = _CENTER
                _ac.font      = _AWAIT_CODE_FONT
                _ac.fill      = _AWAIT_CODE_FILL
                ws.cell(row, S2_QTY, r.item.quantity)
                if S2_NOTES and r.order.private_notes:
                    ws.cell(row, S2_NOTES, r.order.private_notes).alignment = _WRAP

                _style_row(ws, row, COLS, fill=_AWAIT_FILL)
                _ac = ws.cell(row, S2_CHARM)
                _ac.font = _AWAIT_CODE_FONT
                _ac.fill = _AWAIT_CODE_FILL
                _ac.alignment = _CENTER
                if _has_comp_cols:
                    for _mc in (7, 8):
                        ws.cell(row, _mc).fill = _NA_FILL
                        ws.cell(row, _mc).font = _NA_FONT
                if S2_NOTES and r.order.private_notes:
                    ws.cell(row, S2_NOTES).alignment = _WRAP
                ws.cell(row, 1).alignment      = _CENTER
                ws.cell(row, S2_QTY).alignment = _CENTER
                ws.row_dimensions[row].height  = _row_h
                _embed_photo(ws, r.item.photo_bytes, row, S2_PHOTO, _photo_px)
                row += 1

    # ── Column widths ─────────────────────────────────────────────────────
    # EN (10 cols):      # | Photo | Supplier | Stall | ITP | Case | Grip | Phone | Qty | Notes
    # ZH + status (9):   same minus Notes  (Chinese employee status file)
    # ZH compact (7):    # | Photo | Supplier | Stall | ITP | Phone | Qty  (no comp cols)
    if _has_notes_col:
        col_widths = [4, _photo_col_w, 14, 14, 9, 10, 8, 8, 18, 32]
    elif _has_comp_cols:
        col_widths = [4, _photo_col_w, 14, 14, 9, 10, 8, 8, 12]
    else:
        col_widths = [4, _photo_col_w, 14, 9, 12, 18, 8]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Hide the machine-readable round-trip key column so it stays invisible to
    # the human but survives a save/edit/re-upload cycle for the importer.
    if KEY_COL is not None:
        _key_letter = get_column_letter(KEY_COL)
        ws.column_dimensions[_key_letter].width  = 2
        ws.column_dimensions[_key_letter].hidden = True

    ws.freeze_panes = "A5"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{HDR_ROW}:{col_end}{last_data_row}"
def generate_xlsx_simple(
    items: list[ResolvedItem],
    output: Path,
    statuses: dict[tuple[str, str], str] | None = None,
    charm_shops: list[CharmShop] | None = None,
    charm_library: dict[str, CharmLibraryEntry] | None = None,
    charm_images_dir: Path | None = None,
    lang: str = "en",
    show_components: bool | None = None,
) -> None:
    """Generate the simplified single-sheet shopping route workbook.

    Pass ``lang='zh'`` to produce the Chinese-translated variant, which uses
    the same compact layout as the English simple version but with Simplified
    Chinese column headers, section banners, and status dropdown values.

    ``show_components`` overrides whether the per-item Case / Grip status
    columns are rendered.  When ``None`` (default) they follow the language
    convention (shown for English, hidden for the compact Chinese file).  Pass
    ``True`` to force them on — used for the extra Chinese status-tracking
    workbook so employees can record each component's buy status in-file.
    """
    wb = openpyxl.Workbook()
    _sheet_route_simple(
        wb.active, items,
        statuses=statuses,
        charm_shops=charm_shops,
        charm_library=charm_library,
        charm_images_dir=charm_images_dir,
        lang=lang,
        show_components=show_components,
    )
    from openpyxl.worksheet.page import PageMargins
    ws = wb.active
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3,
    )
    wb.save(output)
    log.info("Simple route saved -> %s", output.resolve())
_STATUS_CSS = {
    "Purchased":         "s-bought",
    "Out of Stock":      "s-oos",
    "Out of Production": "s-oop",
    "Pending":           "s-pending",
    "N/A":               "s-na",
    # ZH equivalents
    "已购买": "s-bought",
    "缺货":   "s-oos",
    "停产":   "s-oop",
    "待处理": "s-pending",
    "不适用": "s-na",
}
_ROW_STATUS_CSS = {
    "oop":      "row-oop",
    "oos":      "row-oos",
    "bought":   "row-bought",
    "pending":  "",
}
class RouteOutputLockedError(SystemExit):
    """Raised when a target route file is locked (e.g. open in Excel)."""
def _assert_route_outputs_writable(
    output_path: Path, *, want_chinese: bool, want_html: bool
) -> None:
    """Fail fast — *before* the slow build — if any route output file is locked.

    On Windows a workbook open in Excel holds an exclusive lock, so the later
    ``wb.save()`` raises ``PermissionError`` part-way through the run.  Because
    the full + simple files are written before the Chinese one, a locked
    ``_zh.xlsx`` lets the first two succeed while the Chinese file is left
    untouched — which is exactly why "I regenerate but the Excel never changes".

    We probe every file we are about to write and, if any are locked, abort
    immediately with one clear, actionable message naming the file(s) to close,
    instead of doing all the work and then crashing with a raw traceback.
    """
    targets: list[Path] = [
        output_path,
        output_path.with_stem(output_path.stem + "_simple"),
    ]
    if want_chinese:
        targets.append(output_path.with_stem(output_path.stem + "_zh"))
        targets.append(output_path.with_stem(output_path.stem + "_zh_status"))
    if want_html:
        targets.append(output_path.with_suffix(".html"))
        if want_chinese:
            targets.append(
                output_path.with_stem(output_path.stem + "_zh").with_suffix(".html")
            )

    locked: list[str] = []
    for p in targets:
        if not p.exists():
            continue  # nothing to overwrite — guaranteed writable
        try:
            # Open for read+write WITHOUT truncating; this acquires the same
            # write access wb.save needs, so it fails iff the file is locked.
            with open(p, "r+b"):
                pass
        except PermissionError:
            locked.append(p.name)
        except OSError:
            # Other transient IO errors are not a lock — let the real save report them.
            pass

    if locked:
        names = ", ".join(sorted(set(locked)))
        log.error(
            "Route output file(s) locked: %s — open in Excel or another program. "
            "Close the file(s) and run the generation again.",
            names,
        )
        # A recognisable, single-line message the dashboard can surface verbatim.
        raise RouteOutputLockedError(
            f"ROUTE_OUTPUT_LOCKED: {names} is open in Excel. "
            f"Close it and click Generate again."
        )

def _style_comps(style: str) -> frozenset:
    """Style string -> frozenset of component names ({'case','grip','charm'})."""
    hc, hg, hch = _style_has(style)
    return frozenset(x for x, v in (("case", hc), ("grip", hg), ("charm", hch)) if v)


def _line_disc(item: "OrderItem") -> str:
    """Per-line discriminator (Etsy listing id) so two distinct listings that
    share a 50-char title prefix + component set are never collapsed by dedup."""
    return str(getattr(item, "listing_id", "") or "").strip()


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Generate Etsy shopping-route Excel files from a Unified "
                    "Dashboard --import-json export (or build the ready-to-ship "
                    "report from an edited checklist)."
    )
    ap.add_argument("--project-dir", default="", metavar="DIR",
                    help="Project root (data/, output/, cache/ subdirs).")
    ap.add_argument("--data-dir", default="", metavar="DIR",
                    help="Mutable data directory override (catalog DB/workbook, "
                         "charm manifest and images).")
    ap.add_argument("--catalog", default=CATALOG_FILE, help="Supplier catalog .xlsx")
    ap.add_argument("--output", default=OUTPUT_FILE, help="Output .xlsx path")
    ap.add_argument("--threshold", type=int, default=MATCH_THRESHOLD,
                    help="Fuzzy-match score cutoff 0-100 (default %(default)s)")
    ap.add_argument("--import-json", default="", metavar="FILE",
                    help="Unified Dashboard order export to build the route from.")
    ap.add_argument("--exclude-orders-file", default="", metavar="FILE",
                    help="JSON of (title, order) line items to drop from the route.")
    ap.add_argument("--chinese", action="store_true",
                    help="Also write the Simplified Chinese route files (_zh, "
                         "_zh_status).")
    ap.add_argument("--chinese-exclude-shops", default="", metavar="SHOPS",
                    help="Comma-separated Etsy shop names to omit from Chinese files.")
    ap.add_argument("--charm-images-dir", default="", metavar="DIR",
                    help="Directory of <Charm Code>.png charm photos.")
    ap.add_argument("--no-catalog-update", action="store_true",
                    help="Accepted for compatibility; the catalog is never written.")
    ap.add_argument("--reset", action="store_true",
                    help="Accepted for compatibility; the route is always built "
                         "purely from --import-json (no cache merge).")
    ap.add_argument("--no-charm-manifest", action="store_true",
                    help="Skip writing charm_manifest.json.")
    args = ap.parse_args()

    # ── Path resolution (project-dir organized layout) ────────────────────────
    data_dir_override = (
        Path(args.data_dir).expanduser().resolve()
        if (args.data_dir or "").strip()
        else None
    )
    if args.project_dir:
        proj = Path(args.project_dir).resolve()
        data_dir = data_dir_override or (proj / "data")
        catalog_path = data_dir / "supplier_catalog.xlsx"
        output_path = (
            Path(args.output).expanduser().resolve()
            if (args.output or "").strip() and args.output != OUTPUT_FILE
            else proj / "output" / "shopping_route.xlsx"
        )
        charm_images_dir = (
            Path(args.charm_images_dir).resolve()
            if args.charm_images_dir.strip()
            else data_dir / CHARM_IMAGES_DIR_NAME
        )
    else:
        data_dir = data_dir_override
        catalog_path = (
            data_dir / "supplier_catalog.xlsx"
            if data_dir is not None
            else Path(args.catalog)
        )
        output_path = Path(args.output)
        charm_images_dir = (
            Path(args.charm_images_dir).resolve()
            if args.charm_images_dir.strip()
            else (Path("data") / CHARM_IMAGES_DIR_NAME).resolve()
        )
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        log.warning("Could not create output directory %s: %s", output_path.parent, exc)
    charm_manifest_default_path = (
        ((data_dir_override or (Path(args.project_dir).resolve() / "data"))
         / CHARM_MANIFEST_FILE)
        if args.project_dir.strip() or data_dir_override is not None
        else (catalog_path.parent / CHARM_MANIFEST_FILE)
    ).resolve()

    # ── Build the route from the Unified Dashboard JSON export ────────────────
    import_json_path = (args.import_json or "").strip()
    if not import_json_path:
        ap.error("--import-json is required.")
    if not catalog_path.exists():
        log.error("Catalog not found: %s", catalog_path)
        sys.exit(1)

    _assert_route_outputs_writable(output_path, want_chinese=bool(args.chinese),
                                   want_html=False)
    try:
        charm_images_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass

    # Catalog / charm reference data (read from etsy_orders.db when present).
    catalog = load_catalog(catalog_path)
    charm_shops = load_charm_shops(catalog_path)
    charm_library = load_charm_library(catalog_path)

    all_new_orders = import_orders_from_json(Path(import_json_path))
    total_items = sum(len(o.items) for o in all_new_orders)
    log.info("JSON import: %d order(s), %d line item(s) loaded from '%s'",
             len(all_new_orders), total_items, import_json_path)

    new_resolved = match_items(all_new_orders, catalog, args.threshold,
                               db_path=catalog_path.parent / DB_FILE)
    _apply_import_charm_overrides(new_resolved)
    _apply_import_supplier_overrides(new_resolved)

    # Dedup line items. The route is built purely from the import (reset
    # semantics), so we only guard against duplicates within this run. The key
    # includes the listing id so two distinct listings sharing a 50-char title
    # prefix + component set stay separate (never silently dropped).
    _seen_new: set = set()
    all_resolved: list[ResolvedItem] = []
    for _r in new_resolved:
        _k = (_r.order.order_number, _normalize(_r.item.title)[:50],
              _style_comps(_r.item.style), _line_disc(_r.item))
        if _k in _seen_new:
            continue
        _seen_new.add(_k)
        all_resolved.append(_r)
    duplicates = len(new_resolved) - len(all_resolved)
    if duplicates:
        log.info("Skipped %d duplicate line item(s) within this import", duplicates)
    log.info("Total: %d item(s)  [0 from cache + %d new]",
             len(all_resolved), len(all_resolved))
    if not all_resolved:
        log.error("No order items to write.")
        sys.exit(1)

    # Statuses: preserve any from an existing output workbook, then overlay the
    # authoritative Orders-Dashboard UI cache (route_statuses_cache.json).
    existing_statuses = load_existing_statuses(output_path)
    _ui_overrides = _load_ui_status_cache(output_path.parent)
    if _ui_overrides:
        existing_statuses.update(_ui_overrides)
        log.info("Merged %d UI status override(s)", len(_ui_overrides))

    # The dashboard supplies the authoritative listing photo for every line, so
    # catalog photos only FILL IN items that arrived without an image.
    apply_catalog_photos_to_resolved(all_resolved, catalog_path,
                                     fill_missing_only=True)
    apply_canonical_charm_fields_to_resolved(all_resolved, catalog_path)

    # Drop stale charm_agg=Purchased statuses no longer confirmed by every
    # constituent order (keeps the Chinese charm section truthful).
    _stale_agg = []
    for _sk, _sv in existing_statuses.items():
        if not (len(_sk) == 3 and _sk[2] == "charm_agg" and _sv == "Purchased"):
            continue
        _agg_code = _sk[0]
        _agg_orders = [
            r for r in all_resolved
            if _style_has(r.item.style)[2]
            and (r.supplier.charm_code if r.supplier else "").strip() == _agg_code
        ]
        if not _agg_orders:
            continue
        if not all(
            existing_statuses.get(
                (r.order.order_number, _normalize(r.item.title)[:50], "charm")
            ) == "Purchased"
            for r in _agg_orders
        ):
            _stale_agg.append(_sk)
    for _sk in _stale_agg:
        del existing_statuses[_sk]

    # Optional dashboard exclusion list.
    route_resolved = all_resolved
    if (args.exclude_orders_file or "").strip():
        _excl_path = Path(args.exclude_orders_file.strip())
        if _excl_path.is_file():
            try:
                _excl_pairs = _load_dashboard_route_exclusions(_excl_path)
                if _excl_pairs:
                    route_resolved = [
                        r for r in all_resolved
                        if (_normalize(r.item.title),
                            str(r.order.order_number).strip()) not in _excl_pairs
                    ]
                    log.info("Route: %d of %d item(s) included (%d excluded)",
                             len(route_resolved), len(all_resolved),
                             len(all_resolved) - len(route_resolved))
            except Exception as _excl_err:
                log.warning("Could not apply --exclude-orders-file: %s", _excl_err)

    # ── Write the English workbooks (full + simple) ───────────────────────────
    generate_xlsx(route_resolved, output_path, statuses=existing_statuses,
                  charm_shops=charm_shops, charm_library=charm_library,
                  charm_images_dir=charm_images_dir)
    simple_output_path = output_path.with_stem(output_path.stem + "_simple")
    generate_xlsx_simple(route_resolved, simple_output_path,
                         statuses=existing_statuses, charm_shops=charm_shops,
                         charm_library=charm_library,
                         charm_images_dir=charm_images_dir)

    # ── Write the Simplified Chinese workbooks (_zh, _zh_status) ──────────────
    # The employee edits _zh_status (it carries the hidden round-trip key column)
    # and uploads it back via the dashboard's "Sync purchase status" action, which
    # writes the statuses straight into the database. The separate per-order
    # checklist (_zh_check) and the ready-to-ship report (_zh_ready) it fed are no
    # longer produced — the status workbook is the single source of truth.
    zh_item_count = 0
    if args.chinese:
        excluded_shops = {
            s.strip().lower()
            for s in (args.chinese_exclude_shops or "").split(",") if s.strip()
        }
        zh_items = [r for r in route_resolved
                    if r.order.etsy_shop.lower() not in excluded_shops]
        zh_item_count = len(zh_items)

        zh_path = output_path.with_stem(output_path.stem + "_zh")
        generate_xlsx_simple(zh_items, zh_path, statuses=existing_statuses,
                             charm_shops=charm_shops, charm_library=charm_library,
                             charm_images_dir=charm_images_dir, lang="zh")
        zh_status_path = output_path.with_stem(output_path.stem + "_zh_status")
        generate_xlsx_simple(zh_items, zh_status_path, statuses=existing_statuses,
                             charm_shops=charm_shops, charm_library=charm_library,
                             charm_images_dir=charm_images_dir, lang="zh",
                             show_components=True)

    # ── Charm manifest (consumed by the dashboard charm gallery) ──────────────
    manifest_charm_count = 0
    manifest_written_path = None
    if not args.no_charm_manifest:
        try:
            _route_snap = {
                "order_line_items": len(all_resolved),
                "lines_with_charm_in_order_style": sum(
                    1 for r in all_resolved if _style_has(r.item.style)[2]),
                "matched_lines_with_charm_code": sum(
                    1 for r in all_resolved
                    if r.supplier and (r.supplier.charm_code or "").strip()),
            }
            manifest_charm_count = export_charm_manifest(
                catalog_path, charm_images_dir, charm_manifest_default_path,
                route_snapshot=_route_snap)
            manifest_written_path = charm_manifest_default_path
        except Exception as exc:
            log.warning("Charm manifest (auto) failed: %s", exc)

    # ── Summary ───────────────────────────────────────────────────────────────
    routable_ct = sum(1 for r in route_resolved
                      if r.supplier and (r.supplier.shop_name or r.supplier.stall))
    print(f"\n{'=' * 60}")
    print(f"  [OK]  {routable_ct} item(s) ready  (supplier + location known)")
    print(f"  --->  {output_path.resolve()}  ({len(route_resolved)} items)")
    print(f"  [SIMPLE]  {simple_output_path.resolve()}")
    if manifest_written_path is not None:
        print(f"  [JSON] {manifest_written_path.resolve()}  "
              f"({manifest_charm_count} charm(s))")
    if args.chinese:
        print(f"  [ZH]  {output_path.with_stem(output_path.stem + '_zh').resolve()}"
              f"  ({zh_item_count} items)")
        print(f"  [ZH+STATUS]  "
              f"{output_path.with_stem(output_path.stem + '_zh_status').resolve()}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    try:
        main()
    except RouteOutputLockedError:
        raise
    except PermissionError as _exc:
        _fname = getattr(_exc, "filename", "") or "a route file"
        _name = Path(_fname).name if _fname else "a route file"
        log.error("Could not write %s - it is open in Excel or another program. "
                  "Close it and run the generation again.", _name)
        raise SystemExit(
            f"ROUTE_OUTPUT_LOCKED: {_name} is open in Excel. "
            f"Close it and click Generate again."
        )
