#!/usr/bin/env python3
"""Structural diff between two directories of generated route .xlsx files.

Excel files are zip archives with embedded timestamps + images, so a byte diff
is meaningless. Instead we compare what actually matters for correctness:
  - the set of files produced,
  - each workbook's sheet names,
  - each sheet's dimensions (max row/col),
  - every cell's value (text/number), and
  - the count of embedded images per sheet.

Exit code 0 == identical; 1 == differences (printed).
"""
import sys
from pathlib import Path
import openpyxl

FILES = [
    "shopping_route.xlsx",
    "shopping_route_simple.xlsx",
    "shopping_route_zh.xlsx",
    "shopping_route_zh_status.xlsx",
    "shopping_route_zh_check.xlsx",
]


def workbook_fingerprint(path: Path):
    wb = openpyxl.load_workbook(path, data_only=False)
    fp = {}
    for ws in wb.worksheets:
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[c.coordinate] = c.value
        try:
            n_img = len(ws._images)
        except Exception:
            n_img = 0
        fp[ws.title] = {"dims": (ws.max_row, ws.max_column), "cells": cells, "images": n_img}
    return fp


def diff(a_dir: Path, b_dir: Path) -> int:
    problems = []
    for name in FILES:
        pa, pb = a_dir / name, b_dir / name
        if not pa.exists() or not pb.exists():
            problems.append(f"[{name}] missing: golden={pa.exists()} new={pb.exists()}")
            continue
        fa, fb = workbook_fingerprint(pa), workbook_fingerprint(pb)
        if set(fa) != set(fb):
            problems.append(f"[{name}] sheet names differ: {sorted(fa)} vs {sorted(fb)}")
            continue
        for sheet in fa:
            A, B = fa[sheet], fb[sheet]
            if A["dims"] != B["dims"]:
                problems.append(f"[{name}:{sheet}] dims {A['dims']} vs {B['dims']}")
            if A["images"] != B["images"]:
                problems.append(f"[{name}:{sheet}] image count {A['images']} vs {B['images']}")
            keys = set(A["cells"]) | set(B["cells"])
            cell_diffs = []
            for k in sorted(keys):
                va, vb = A["cells"].get(k), B["cells"].get(k)
                if va != vb:
                    cell_diffs.append(f"    {k}: {va!r} != {vb!r}")
            if cell_diffs:
                problems.append(f"[{name}:{sheet}] {len(cell_diffs)} cell diff(s):\n" + "\n".join(cell_diffs[:25]))
    if problems:
        print("DIFFERENCES FOUND:")
        for p in problems:
            print(" -", p)
        return 1
    print("IDENTICAL: all 5 workbooks match (sheets, dims, every cell value, image counts).")
    return 0


if __name__ == "__main__":
    sys.exit(diff(Path(sys.argv[1]), Path(sys.argv[2])))
